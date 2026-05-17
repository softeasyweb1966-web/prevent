from __future__ import annotations

import argparse
import os
import re
import shutil
import unicodedata
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path

import openpyxl
from sqlalchemy import inspect, text

from app import create_app
from app.models import ClienteComercial, ComercialCatalogoItem, Vendedor, db


EXPECTED_HEADERS = [
    'FormaPago',
    'Precio',
    'Nombre del Producto o Servicio',
    'Nombre del Acuerdo Comercial',
]


@dataclass
class ClienteImportRow:
    razon_social: str
    condicion_comercial: str
    requiere_factura: bool
    examenes_convenidos: list[str]
    formas_pago: list[str]
    filas_origen: int


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            'Importa clientes reales desde CLIENTES-2026.xlsx, elimina los clientes '
            'comerciales actuales y asigna todos al vendedor indicado.'
        )
    )
    parser.add_argument(
        'archivo',
        nargs='?',
        default='CLIENTES-2026.xlsx',
        help='Ruta del archivo Excel fuente.',
    )
    parser.add_argument(
        '--sheet',
        default='clientes',
        help='Nombre de la hoja origen. La búsqueda no distingue mayúsculas/minúsculas.',
    )
    parser.add_argument(
        '--vendedor',
        default='Preventsalud',
        help='Nombre de referencia del vendedor al que se asignarán todos los clientes.',
    )
    parser.add_argument(
        '--database-url',
        default=os.environ.get('DATABASE_URL', ''),
        help='URL de PostgreSQL. Si no se envía, usa DATABASE_URL.',
    )
    parser.add_argument(
        '--apply',
        action='store_true',
        help='Ejecuta el import real. Sin esta bandera solo muestra validaciones y resumen.',
    )
    return parser.parse_args()


def _normalize_text(value: object) -> str:
    if value is None:
        return ''
    return ' '.join(str(value).replace('\xa0', ' ').strip().split())


def _normalize_match(value: object) -> str:
    text = _normalize_text(value)
    text = unicodedata.normalize('NFKD', text)
    text = ''.join(ch for ch in text if not unicodedata.combining(ch))
    text = text.lower()
    text = re.sub(r'[^a-z0-9]+', ' ', text)
    return re.sub(r'\s+', ' ', text).strip()


def _find_sheet(workbook, target_name: str):
    normalized_target = _normalize_match(target_name)
    for sheet_name in workbook.sheetnames:
        if _normalize_match(sheet_name) == normalized_target:
            return workbook[sheet_name]
    raise ValueError(f'No existe la hoja "{target_name}" en el archivo. Hojas disponibles: {", ".join(workbook.sheetnames)}')


def _load_rows(path: Path, sheet_name: str) -> list[tuple]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    worksheet = _find_sheet(workbook, sheet_name)
    return list(worksheet.iter_rows(values_only=True))


def _build_grouped_clients(rows: list[tuple]) -> list[ClienteImportRow]:
    if not rows:
        raise ValueError('El archivo está vacío.')

    headers = [_normalize_text(value) for value in rows[0]]
    if headers[:len(EXPECTED_HEADERS)] != EXPECTED_HEADERS:
        raise ValueError(
            'La hoja no tiene la estructura esperada. '
            f'Se esperaban las columnas: {", ".join(EXPECTED_HEADERS)}'
        )

    index = {header: idx for idx, header in enumerate(headers)}
    grouped = {}

    for row_index, row in enumerate(rows[1:], start=2):
        if not row or all(value is None for value in row):
            continue

        razon_social = _normalize_text(row[index['Nombre del Acuerdo Comercial']])
        forma_pago = _normalize_text(row[index['FormaPago']]).upper()
        examen = _normalize_text(row[index['Nombre del Producto o Servicio']])
        if not razon_social or not examen:
            continue

        key = _normalize_match(razon_social)
        bucket = grouped.setdefault(key, {
            'razon_social': razon_social[:200],
            'formas_pago': set(),
            'examenes': {},
            'filas': 0,
        })
        bucket['formas_pago'].add(forma_pago)
        bucket['examenes'].setdefault(_normalize_match(examen), examen[:300])
        bucket['filas'] += 1

    clientes = []
    for bucket in grouped.values():
        formas = sorted(value for value in bucket['formas_pago'] if value)
        if formas == ['CONTADO']:
            condicion = 'EFECTIVO'
            requiere_factura = False
        elif formas == ['CRÉDITO'] or formas == ['CREDITO']:
            condicion = 'CREDITO'
            requiere_factura = True
        else:
            condicion = 'MIXTO'
            requiere_factura = True

        examenes = sorted(bucket['examenes'].values(), key=lambda value: _normalize_match(value))
        clientes.append(ClienteImportRow(
            razon_social=bucket['razon_social'],
            condicion_comercial=condicion,
            requiere_factura=requiere_factura,
            examenes_convenidos=examenes,
            formas_pago=formas,
            filas_origen=bucket['filas'],
        ))

    clientes.sort(key=lambda item: _normalize_match(item.razon_social))
    return clientes


def _resolve_vendedor(nombre_objetivo: str) -> Vendedor:
    normalized_target = _normalize_match(nombre_objetivo)
    candidates = []
    for vendedor in Vendedor.query.order_by(Vendedor.nombre.asc()).all():
        normalized_name = _normalize_match(vendedor.nombre)
        if not normalized_name:
            continue
        if normalized_target == normalized_name or normalized_target in normalized_name or normalized_name in normalized_target:
            candidates.append(vendedor)

    if not candidates:
        raise ValueError(f'No existe un vendedor que coincida con "{nombre_objetivo}".')
    if len(candidates) > 1:
        raise ValueError(
            'La referencia del vendedor es ambigua. Coincidencias: '
            + ', '.join(vendedor.nombre for vendedor in candidates)
        )
    return candidates[0]


def _validate_catalog(clientes: list[ClienteImportRow]) -> list[str]:
    catalogo = ComercialCatalogoItem.query.filter_by(tipo_item='EXAMEN', activo=True).all()
    catalogo_por_clave = {
        _normalize_match(item.nombre): item
        for item in catalogo
    }

    faltantes = set()
    for cliente in clientes:
        for examen in cliente.examenes_convenidos:
            item = catalogo_por_clave.get(_normalize_match(examen))
            if item is None or item.clasificacion_completa is not True:
                faltantes.add(examen)
    return sorted(faltantes, key=_normalize_match)


def _clear_client_uploads() -> None:
    uploads_root = Path(__file__).resolve().parent / 'app' / 'uploads' / 'comercial' / 'clientes'
    if uploads_root.exists():
        shutil.rmtree(uploads_root)
    uploads_root.mkdir(parents=True, exist_ok=True)


def _nullify_atenciones_dia_cliente_links() -> int:
    inspector = inspect(db.engine)
    if 'atenciones_dia_detalle' not in inspector.get_table_names():
        return 0

    columnas = {column['name'] for column in inspector.get_columns('atenciones_dia_detalle')}
    if 'cliente_id' not in columnas:
        return 0

    result = db.session.execute(text('UPDATE atenciones_dia_detalle SET cliente_id = NULL WHERE cliente_id IS NOT NULL'))
    return int(result.rowcount or 0)


def _bulk_delete_if_table_exists(table_name: str, where_clause: str | None = None) -> int:
    inspector = inspect(db.engine)
    if table_name not in inspector.get_table_names():
        return 0

    statement = f'DELETE FROM {table_name}'
    if where_clause:
        statement = f'{statement} WHERE {where_clause}'
    result = db.session.execute(text(statement))
    return int(result.rowcount or 0)


def _import_clientes(clientes: list[ClienteImportRow], vendedor: Vendedor) -> dict[str, int]:
    total_actuales = ClienteComercial.query.count()
    enlaces_limpiados = _nullify_atenciones_dia_cliente_links()
    eliminados_pagos = _bulk_delete_if_table_exists('clientes_seguimiento_pagos')
    eliminados_documentos = _bulk_delete_if_table_exists('clientes_seguimiento_documentos')
    eliminados_atenciones_detalle = _bulk_delete_if_table_exists('clientes_atenciones_detalle')
    eliminados_atenciones = _bulk_delete_if_table_exists('clientes_atenciones')
    eliminados_tarifas = _bulk_delete_if_table_exists('clientes_comerciales_tarifas')
    eliminados_adjuntos = _bulk_delete_if_table_exists('clientes_comerciales_adjuntos')
    eliminados_clientes = _bulk_delete_if_table_exists('clientes_comerciales')

    nuevos = []
    for cliente in clientes:
        nuevos.append(ClienteComercial(
            vendedor_id=vendedor.id,
            razon_social=cliente.razon_social,
            nombre_comercial=None,
            nit=None,
            ciudad=None,
            direccion=None,
            telefono_empresa=None,
            email_empresa=None,
            contacto_principal=None,
            cargo_contacto_principal=None,
            celular_contacto_principal=None,
            email_contacto_principal=None,
            contacto_facturacion=None,
            cargo_contacto_facturacion=None,
            celular_facturacion=None,
            email_facturacion=None,
            medio_autorizacion='WHATSAPP',
            puntos_atencion_recepcion=None,
            estado_cliente='ACTIVO',
            condicion_comercial=cliente.condicion_comercial,
            requiere_factura=cliente.requiere_factura,
            fechas_facturacion=None,
            fecha_solicitud_factura=None,
            examenes_convenidos='\n'.join(cliente.examenes_convenidos),
            servicios_convenidos=None,
            tarifas_convenidas=None,
            documentos_legales_completos=False,
            documentos_legales_detalle=None,
            pagare_firmado=False,
            pagare_detalle=None,
            observaciones=(
                f'Importado desde CLIENTES-2026.xlsx | Formas de pago detectadas: '
                f'{", ".join(cliente.formas_pago) or "N/D"} | Exámenes: {len(cliente.examenes_convenidos)}'
            ),
            activo=True,
        ))

    db.session.add_all(nuevos)
    db.session.commit()
    _clear_client_uploads()

    return {
        'clientes_previos_detectados': total_actuales,
        'clientes_eliminados': eliminados_clientes,
        'adjuntos_eliminados': eliminados_adjuntos,
        'tarifas_eliminadas': eliminados_tarifas,
        'atenciones_eliminadas': eliminados_atenciones,
        'atenciones_detalle_eliminadas': eliminados_atenciones_detalle,
        'documentos_eliminados': eliminados_documentos,
        'pagos_eliminados': eliminados_pagos,
        'clientes_insertados': len(nuevos),
        'atenciones_dia_desvinculadas': enlaces_limpiados,
    }


def main() -> int:
    args = _parse_args()
    archivo = Path(args.archivo)
    if not archivo.exists():
        raise SystemExit(f'No existe el archivo fuente: {archivo}')
    if not args.database_url:
        raise SystemExit('Debes enviar --database-url o definir DATABASE_URL.')

    os.environ['DATABASE_URL'] = args.database_url
    clientes = _build_grouped_clients(_load_rows(archivo, args.sheet))
    total_examenes_unicos = len({
        _normalize_match(examen)
        for cliente in clientes
        for examen in cliente.examenes_convenidos
    })

    app = create_app()
    with app.app_context():
        vendedor = _resolve_vendedor(args.vendedor)
        faltantes = _validate_catalog(clientes)
        if faltantes:
            print('Se encontraron exámenes faltantes o no utilizables en convenios:')
            for examen in faltantes:
                print(f'- {examen}')
            raise SystemExit(2)

        print(f'Archivo: {archivo.name}')
        print(f'Hoja: {args.sheet}')
        print(f'Vendedor destino: {vendedor.nombre} (id={vendedor.id})')
        print(f'Clientes detectados: {len(clientes)}')
        print(f'Exámenes únicos detectados: {total_examenes_unicos}')
        print(f'Clientes actuales en BD: {ClienteComercial.query.count()}')

        if not args.apply:
            print('Validación completada. Usa --apply para ejecutar el import real.')
            return 0

        try:
            resumen = _import_clientes(clientes, vendedor)
        except Exception:
            db.session.rollback()
            raise

    print('Importación completada correctamente.')
    for key, value in resumen.items():
        print(f'{key}: {value}')
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
