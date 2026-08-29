"""Endpoints para cargue diario de atenciones desde Excel."""

from __future__ import annotations

import io
import json
import logging
import os
import re
import types
import unicodedata
import uuid
import zipfile
from collections import defaultdict
from datetime import datetime
from decimal import Decimal, InvalidOperation
from xml.etree import ElementTree as ET

from flask import current_app, jsonify, request, send_file
from flask_login import current_user, login_required
from sqlalchemy import BigInteger, String, and_, cast, func, or_, text
from sqlalchemy.exc import OperationalError, ProgrammingError
from werkzeug.utils import secure_filename

from app.models import (
    AtencionDiaDetalle,
    CarteraPrefactura,
    CargueAtencionDia,
    ClienteComercial,
    ComercialCatalogoItem,
    OrdenServicioCaja,
    OrdenServicioCajaAdjunto,
    PrefacturaComercial,
    PrefacturaComercialDetalle,
    Vendedor,
    db,
)
from app.routes import comercial_bp
from app.security import get_permission_names_for_user

logger = logging.getLogger(__name__)


def _mensaje_error_estado_gestion_atenciones(exc: Exception) -> str | None:
    detalle = f'{exc} {getattr(exc, "orig", "")}'.lower()
    if 'estado_gestion' not in detalle:
        return None
    return (
        'La base de datos aun no tiene el nuevo campo "estado_gestion" de Atenciones. '
        'Aplica la migracion pendiente con "flask db upgrade".'
    )


def _mensaje_error_esquema_prefacturas(exc: Exception) -> str | None:
    detalle = f'{exc} {getattr(exc, "orig", "")}'.lower()
    if 'prefacturas_comerciales' not in detalle:
        return None
    columnas_clave = (
        'origen',
        'fecha_programada',
        'bloqueada_por_pago',
        'fecha_bloqueo_pago',
    )
    if not any(columna in detalle for columna in columnas_clave):
        return None
    return (
        'La base de datos aun no tiene los nuevos campos de Prefacturas Comerciales. '
        'La sabana se genero, pero no se pudo sincronizar el registro en BD. '
        'Aplica la migracion pendiente con "flask db upgrade".'
    )


def _columnas_periodo_cargue_disponibles() -> bool:
    try:
        rows = db.session.execute(text("""
            select column_name
            from information_schema.columns
            where table_schema = 'public'
              and table_name = 'cargue_atenciones_dia'
              and column_name in ('periodo_desde', 'periodo_hasta')
        """)).fetchall()
    except Exception:
        return True
    columnas = {row[0] for row in rows}
    return {'periodo_desde', 'periodo_hasta'}.issubset(columnas)


def _columnas_prefactura_avanzadas_disponibles() -> bool:
    try:
        rows = db.session.execute(text("""
            select column_name
            from information_schema.columns
            where table_schema = 'public'
              and table_name = 'prefacturas_comerciales'
              and column_name in ('origen', 'fecha_programada', 'bloqueada_por_pago', 'fecha_bloqueo_pago')
        """)).fetchall()
    except Exception:
        return True
    columnas = {row[0] for row in rows}
    return {'origen', 'fecha_programada', 'bloqueada_por_pago', 'fecha_bloqueo_pago'}.issubset(columnas)


def _detalle_prefactura_manual_disponible() -> bool:
    try:
        rows = db.session.execute(text("""
            select table_name
            from information_schema.tables
            where table_schema = 'public'
              and table_name = 'prefacturas_comerciales_detalle'
        """)).fetchall()
    except Exception:
        return True
    return bool(rows) and _columnas_prefactura_avanzadas_disponibles()


def _crear_cargue_atenciones(nombre_archivo, total_filas, usuario_id, periodo_desde, periodo_hasta):
    if _columnas_periodo_cargue_disponibles():
        cargue = CargueAtencionDia(
            nombre_archivo=nombre_archivo,
            periodo_desde=periodo_desde,
            periodo_hasta=periodo_hasta,
            total_filas=total_filas,
            usuario_id=usuario_id,
        )
        db.session.add(cargue)
        db.session.flush()
        return cargue, False

    created_at = datetime.utcnow()
    row = db.session.execute(text("""
        insert into cargue_atenciones_dia
            (nombre_archivo, total_filas, filas_importadas, filas_duplicadas, filas_error, usuario_id, created_at)
        values
            (:nombre_archivo, :total_filas, 0, 0, 0, :usuario_id, :created_at)
        returning id
    """), {
        'nombre_archivo': nombre_archivo,
        'total_filas': total_filas,
        'usuario_id': usuario_id,
        'created_at': created_at,
    }).first()
    cargue = types.SimpleNamespace(
        id=row.id,
        nombre_archivo=nombre_archivo,
        periodo_desde=None,
        periodo_hasta=None,
        total_filas=total_filas,
        filas_importadas=0,
        filas_duplicadas=0,
        filas_error=0,
        created_at=created_at,
    )
    return cargue, True


def _actualizar_resumen_cargue(cargue, importadas, duplicadas, errores, legacy_mode=False):
    cargue.filas_importadas = importadas
    cargue.filas_duplicadas = duplicadas
    cargue.filas_error = errores
    if legacy_mode:
        db.session.execute(text("""
            update cargue_atenciones_dia
            set filas_importadas = :importadas,
                filas_duplicadas = :duplicadas,
                filas_error = :errores
            where id = :cargue_id
        """), {
            'importadas': importadas,
            'duplicadas': duplicadas,
            'errores': errores,
            'cargue_id': cargue.id,
        })

COLUMNAS_ESPERADAS = [
    'NÃâÃÂ°. Orden Servicio',
    'NÃâÃÂ°. Factura',
    'Fecha de Factura',
    'Precio',
    'FormaPago',
    'Nombre del Producto o Servicio',
    'NÃâÃÂ°. de IdentificaciÃÆÃÂ³n',
    'Nombre del Paciente',
    'Nombre del Acuerdo Comercial',
    'Empresa en MisiÃÆÃÂ³n',
    'Sede',
    'Nombre del Vendedor',
    'Fecha de CreaciÃÆÃÂ³n Orden Servicio',
    'Usuario de CreaciÃÆÃÂ³n Orden Servicio',
    'Estado de la Orden Servicio',
    'Fecha de AnulaciÃÆÃÂ³n Orden Servicio',
]

COL_ORDEN_SERVICIO = 'Nro Orden Servicio'
COL_FACTURA = 'Nro Factura'
COL_FECHA_FACTURA = 'Fecha de Factura'
COL_PRECIO = 'Precio'
COL_FORMA_PAGO = 'FormaPago'
COL_PRODUCTO_SERVICIO = 'Nombre del Producto o Servicio'
COL_IDENTIFICACION = 'Nro de Identificacion'
COL_PACIENTE = 'Nombre del Paciente'
COL_ACUERDO_COMERCIAL = 'Nombre del Acuerdo Comercial'
COL_EMPRESA_MISION = 'Empresa en Mision'
COL_SEDE = 'Sede'
COL_VENDEDOR = 'Nombre del Vendedor'
COL_FECHA_CREACION_ORDEN = 'Fecha de Creacion Orden Servicio'
COL_USUARIO_CREACION = 'Usuario de Creacion Orden Servicio'
COL_ESTADO_ORDEN = 'Estado de la Orden Servicio'
COL_FECHA_ANULACION = 'Fecha de Anulacion Orden Servicio'

COLUMNAS_ESPERADAS_CANONICAS = {
    COL_ORDEN_SERVICIO: ['Nro Orden Servicio', 'Nro. Orden Servicio', 'N°. Orden Servicio', 'Nº. Orden Servicio', 'No. Orden Servicio'],
    COL_FACTURA: ['Nro Factura', 'Nro. Factura', 'N°. Factura', 'Nº. Factura', 'No. Factura'],
    COL_FECHA_FACTURA: ['Fecha de Factura'],
    COL_PRECIO: ['Precio'],
    COL_FORMA_PAGO: ['FormaPago', 'Forma Pago'],
    COL_PRODUCTO_SERVICIO: ['Nombre del Producto o Servicio'],
    COL_IDENTIFICACION: ['Nro de Identificacion', 'Nro. de Identificacion', 'N°. de Identificacion', 'N°. de Identificación', 'Nº. de Identificacion', 'Nº. de Identificación', 'No. de Identificacion'],
    COL_PACIENTE: ['Nombre del Paciente'],
    COL_ACUERDO_COMERCIAL: ['Nombre del Acuerdo Comercial'],
    COL_EMPRESA_MISION: ['Empresa en Mision', 'Empresa en Misión'],
    COL_SEDE: ['Sede'],
    COL_VENDEDOR: ['Nombre del Vendedor'],
    COL_FECHA_CREACION_ORDEN: ['Fecha de Creacion Orden Servicio', 'Fecha de Creación Orden Servicio'],
    COL_USUARIO_CREACION: ['Usuario de Creacion Orden Servicio', 'Usuario de Creación Orden Servicio'],
    COL_ESTADO_ORDEN: ['Estado de la Orden Servicio'],
    COL_FECHA_ANULACION: ['Fecha de Anulacion Orden Servicio', 'Fecha de Anulación Orden Servicio'],
}

COLUMNAS_ESPERADAS_LEGACY = {
    COL_ORDEN_SERVICIO: COLUMNAS_ESPERADAS[0],
    COL_FACTURA: COLUMNAS_ESPERADAS[1],
    COL_FECHA_FACTURA: COLUMNAS_ESPERADAS[2],
    COL_PRECIO: COLUMNAS_ESPERADAS[3],
    COL_FORMA_PAGO: COLUMNAS_ESPERADAS[4],
    COL_PRODUCTO_SERVICIO: COLUMNAS_ESPERADAS[5],
    COL_IDENTIFICACION: COLUMNAS_ESPERADAS[6],
    COL_PACIENTE: COLUMNAS_ESPERADAS[7],
    COL_ACUERDO_COMERCIAL: COLUMNAS_ESPERADAS[8],
    COL_EMPRESA_MISION: COLUMNAS_ESPERADAS[9],
    COL_SEDE: COLUMNAS_ESPERADAS[10],
    COL_VENDEDOR: COLUMNAS_ESPERADAS[11],
    COL_FECHA_CREACION_ORDEN: COLUMNAS_ESPERADAS[12],
    COL_USUARIO_CREACION: COLUMNAS_ESPERADAS[13],
    COL_ESTADO_ORDEN: COLUMNAS_ESPERADAS[14],
    COL_FECHA_ANULACION: COLUMNAS_ESPERADAS[15],
}

PERMISO_CARGUE_ATENCIONES = 'comercial_atenciones_create'
PERMISO_CONSULTA_ATENCIONES = 'comercial_atenciones_read'
PERMISO_EDICION_ATENCIONES = 'comercial_atenciones_update'
PERMISO_ELIMINACION_ATENCIONES = 'comercial_atenciones_delete'
NOMBRE_CARGUE_MANUAL = 'REGISTRO_MANUAL_ATENCIONES'
ESTADOS_GESTION_ATENCION = {'CARGADA', 'CON_FACTURA', 'TERMINADA', 'ANULADA'}

_FECHA_RE = re.compile(r'(\d{1,2})/(\d{1,2})/(\d{4})\s+(\d{1,2}):(\d{2})')
_XML_NS = {'a': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
_VALORES_EMPRESA_GENERICOS = {
    'principal',
    'particular',
    'ninguna',
    'na',
    'n a',
}


def _is_admin_user():
    return bool(getattr(current_user, 'is_easy', False)) or getattr(getattr(current_user, 'role', None), 'nombre', None) == 'Administrador'


def _get_current_permission_names():
    if _is_admin_user():
        return {'*'}
    return get_permission_names_for_user(current_user)


def _require_commercial_permission(permission_name):
    if '*' in _get_current_permission_names():
        return
    if permission_name not in _get_current_permission_names():
        raise PermissionError('No tienes permiso para realizar esta acciÃÆÃÂ³n en atenciones comerciales')


def _normalizar(valor):
    if valor is None:
        return None
    texto = str(valor).replace('\xa0', ' ').strip()
    return texto or None


def _normalizar_etiqueta(valor):
    texto = _normalizar(valor) or ''
    texto = unicodedata.normalize('NFKD', texto)
    texto = ''.join(ch for ch in texto if not unicodedata.combining(ch))
    texto = texto.lower().replace('nÃâÃÂ°.', 'nro').replace('nÃâÃÂ°', 'nro').replace('nÃâÃÂº', 'nro')
    texto = re.sub(r'[^a-z0-9]+', ' ', texto)
    return re.sub(r'\s+', ' ', texto).strip()


def _normalizar_match(valor):
    texto = _normalizar(valor) or ''
    texto = unicodedata.normalize('NFKD', texto)
    texto = ''.join(ch for ch in texto if not unicodedata.combining(ch))
    texto = texto.lower()
    texto = re.sub(r'[^a-z0-9]+', ' ', texto)
    return re.sub(r'\s+', ' ', texto).strip()


def _normalizar_encabezado_atencion(valor):
    texto = _normalizar(valor) or ''
    reemplazos = {
        'N°': 'Nro ',
        'Nº': 'Nro ',
        'No.': 'Nro ',
        'No ': 'Nro ',
    }
    for origen, destino in reemplazos.items():
        texto = texto.replace(origen, destino)
    normalizado = _normalizar_match(texto)
    normalizado = re.sub(r'^n (?=orden servicio\b)', 'nro ', normalizado)
    normalizado = re.sub(r'^n (?=factura\b)', 'nro ', normalizado)
    normalizado = re.sub(r'^n (?=de identific)', 'nro ', normalizado)
    normalizado = normalizado.replace('identificaci n', 'identificacion')
    normalizado = normalizado.replace('misi n', 'mision')
    normalizado = normalizado.replace('creaci n', 'creacion')
    normalizado = normalizado.replace('anulaci n', 'anulacion')
    return normalizado


def _parse_fecha(valor):
    """Parsea fechas en formato dd/mm/yyyy hh:mm a./p. m."""
    if valor is None:
        return None
    if isinstance(valor, datetime):
        return valor

    texto = _normalizar(valor)
    if not texto:
        return None

    for fmt in ('%Y-%m-%d', '%Y-%m-%d %H:%M:%S', '%d/%m/%Y', '%d/%m/%Y %H:%M'):
        try:
            return datetime.strptime(texto, fmt)
        except ValueError:
            pass

    match = _FECHA_RE.search(texto)
    if not match:
        return None

    dia, mes, anio, hora, minuto = (int(x) for x in match.groups())
    texto_lower = texto.lower()
    if 'p' in texto_lower and hora != 12:
        hora += 12
    elif 'a' in texto_lower and hora == 12:
        hora = 0

    try:
        return datetime(anio, mes, dia, hora, minuto)
    except ValueError:
        return None


def _parse_precio(valor):
    if valor is None:
        return None
    if isinstance(valor, Decimal):
        return valor
    if isinstance(valor, (int, float)):
        return Decimal(str(valor))

    texto = _normalizar(valor)
    if not texto:
        return None

    texto = texto.replace('$', '').replace(' ', '')
    if ',' in texto and '.' in texto:
        if texto.rfind(',') > texto.rfind('.'):
            texto = texto.replace('.', '').replace(',', '.')
        else:
            texto = texto.replace(',', '')
    elif ',' in texto:
        texto = texto.replace('.', '').replace(',', '.')

    try:
        return Decimal(texto)
    except InvalidOperation:
        return None


def _cargar_rows_openpyxl(contenido):
    import openpyxl

    workbook = openpyxl.load_workbook(io.BytesIO(contenido), read_only=True, data_only=True)
    worksheet = workbook.active
    return [list(row) for row in worksheet.iter_rows(values_only=True)]


def _cell_value_xml(cell, shared_strings):
    cell_type = cell.attrib.get('t')
    if cell_type == 'inlineStr':
        return ''.join(node.text or '' for node in cell.findall('.//a:t', _XML_NS))

    value_node = cell.find('a:v', _XML_NS)
    if value_node is None:
        return None

    raw_value = value_node.text
    if cell_type == 's':
        try:
            return shared_strings[int(raw_value)]
        except Exception:
            return raw_value
    return raw_value


def _columna_excel_a_indice(ref):
    indice = 0
    for caracter in ref:
        if not caracter.isalpha():
            break
        indice = (indice * 26) + (ord(caracter.upper()) - 64)
    return indice - 1


def _cargar_rows_xlsx_xml(contenido):
    with zipfile.ZipFile(io.BytesIO(contenido)) as zipped:
        shared_strings = []
        if 'xl/sharedStrings.xml' in zipped.namelist():
            root = ET.fromstring(zipped.read('xl/sharedStrings.xml'))
            for node in root.findall('a:si', _XML_NS):
                shared_strings.append(''.join(text_node.text or '' for text_node in node.findall('.//a:t', _XML_NS)))

        root = ET.fromstring(zipped.read('xl/worksheets/sheet1.xml'))
        rows = []
        for row in root.findall('.//a:sheetData/a:row', _XML_NS):
            values_by_index = {}
            max_index = -1
            for cell in row.findall('a:c', _XML_NS):
                reference = cell.attrib.get('r', '')
                column_index = _columna_excel_a_indice(reference)
                max_index = max(max_index, column_index)
                values_by_index[column_index] = _cell_value_xml(cell, shared_strings)

            if max_index < 0:
                rows.append([])
                continue

            dense_row = [None] * (max_index + 1)
            for column_index, value in values_by_index.items():
                dense_row[column_index] = value
            rows.append(dense_row)

    return rows


def _leer_excel_atenciones(contenido, nombre_archivo):
    if not str(nombre_archivo or '').lower().endswith('.xlsx'):
        raise ValueError('Solo se aceptan archivos .xlsx con la estructura de Estado-de-las-Atenciones.xlsx')

    try:
        return _cargar_rows_openpyxl(contenido)
    except ImportError:
        logger.warning('openpyxl no estÃÆÃÂ¡ disponible; se usarÃÆÃÂ¡ parser XML para %s', nombre_archivo)
    except Exception as exc:
        logger.warning('openpyxl no pudo leer %s (%s). Se intentarÃÆÃÂ¡ parser XML.', nombre_archivo, exc)

    try:
        return _cargar_rows_xlsx_xml(contenido)
    except Exception as exc:
        logger.error('Error leyendo Excel de atenciones con parser XML: %s', exc)
        raise ValueError(f'No se pudo leer el archivo Excel: {exc}') from exc


def _extraer_registros_excel(filas):
    if not filas:
        raise ValueError('El archivo estÃÆÃÂ¡ vacÃÆÃÂ­o')

    encabezados = [str(valor).strip() if valor is not None else '' for valor in filas[0]]
    headers_norm = {
        _normalizar_encabezado_atencion(valor): index
        for index, valor in enumerate(encabezados)
        if _normalizar_encabezado_atencion(valor)
    }

    required_map = {}
    faltantes = []
    for columna_canonica, aliases in COLUMNAS_ESPERADAS_CANONICAS.items():
        encontrado = None
        for alias in aliases:
            alias_norm = _normalizar_encabezado_atencion(alias)
            if alias_norm in headers_norm:
                encontrado = alias_norm
                break
        if encontrado is None:
            faltantes.append(columna_canonica)
        else:
            required_map[encontrado] = columna_canonica
    if faltantes:
        raise ValueError(
            f'Columnas requeridas no encontradas: {", ".join(faltantes)}. '
            'Verifique que el archivo tenga el mismo formato de Estado-de-las-Atenciones.xlsx.'
        )

    registros = []
    for fila in filas[1:]:
        if not fila or all(valor in (None, '') for valor in fila):
            continue

        registro = {}
        for normalized_name, original_name in required_map.items():
            index = headers_norm[normalized_name]
            valor = fila[index] if index < len(fila) else None
            registro[original_name] = valor
            legacy_name = COLUMNAS_ESPERADAS_LEGACY.get(original_name)
            if legacy_name:
                registro[legacy_name] = valor
        registros.append(registro)

    return registros


def _resolver_vendedor_usuario_actual():
    if _is_admin_user():
        return None

    normalized_candidates = [
        _normalizar_match(getattr(current_user, 'email', None)),
        _normalizar_match(getattr(current_user, 'usuario', None)),
        _normalizar_match(getattr(current_user, 'nombre_completo', None)),
    ]
    normalized_candidates = [value for value in normalized_candidates if value]
    if not normalized_candidates:
        return None

    vendedores = Vendedor.query.all()
    for vendedor in vendedores:
        vendor_candidates = {
            _normalizar_match(vendedor.nombre),
            _normalizar_match(vendedor.email),
            _normalizar_match(vendedor.documento),
        }
        vendor_candidates.discard('')
        if any(candidate in vendor_candidates for candidate in normalized_candidates):
            return vendedor
    return None


def _construir_lookup_clientes():
    lookup = {}
    clientes = ClienteComercial.query.all()
    for cliente in clientes:
        for value in (cliente.razon_social, cliente.nombre_comercial, cliente.nit):
            normalized = _normalizar_match(value)
            if normalized and normalized not in lookup:
                lookup[normalized] = cliente
    return lookup


def _construir_lookup_vendedores():
    lookup = {}
    vendedores = Vendedor.query.all()
    for vendedor in vendedores:
        for value in (vendedor.nombre, vendedor.email, vendedor.documento):
            normalized = _normalizar_match(value)
            if normalized and normalized not in lookup:
                lookup[normalized] = vendedor
    return lookup


def _clientes_visibles_query(vendedor_scope):
    query = ClienteComercial.query
    if not _is_admin_user():
        if vendedor_scope is None:
            return query.filter(ClienteComercial.id == -1)
        query = query.filter(ClienteComercial.vendedor_id == vendedor_scope.id)
    return query


def _cliente_desde_registro(registro, clientes_lookup):
    for campo in ('Nombre del Acuerdo Comercial', 'Empresa en MisiÃÆÃÂ³n'):
        normalized = _normalizar_match(registro.get(campo))
        if not normalized or normalized in _VALORES_EMPRESA_GENERICOS:
            continue
        cliente = clientes_lookup.get(normalized)
        if cliente is not None:
            return cliente
    return None


def _vendedor_desde_registro(registro, vendedores_lookup):
    normalized = _normalizar_match(registro.get('Nombre del Vendedor'))
    if not normalized:
        return None
    return vendedores_lookup.get(normalized)


def _scope_descriptor(vendedor):
    if _is_admin_user():
        return 'admin'
    if vendedor is None:
        return 'sin_vendedor_asociado'
    return 'vendedor'


def _asegurar_acceso_registro_atencion(registro, vendedor_scope):
    if _is_admin_user():
        return
    if vendedor_scope is None or registro.vendedor_id != vendedor_scope.id:
        raise PermissionError('No tienes permiso para acceder a esta atencion comercial')


def _construir_lookup_clientes_visibles(vendedor_scope):
    lookup = {}
    clientes = _clientes_visibles_query(vendedor_scope).all()
    for cliente in clientes:
        for value in (cliente.razon_social, cliente.nombre_comercial, cliente.nit):
            normalized = _normalizar_match(value)
            if normalized and normalized not in lookup:
                lookup[normalized] = cliente
    return lookup


def _serialize_atencion_dia(registro):
    return {
        'id': registro.id,
        'cargue_id': registro.cargue_id,
        'cliente_id': registro.cliente_id,
        'cliente_nombre': registro.cliente.razon_social if registro.cliente else None,
        'vendedor_id': registro.vendedor_id,
        'vendedor_responsable': registro.vendedor.nombre if registro.vendedor else None,
        'nro_orden': registro.nro_orden,
        'nro_factura': registro.nro_factura,
        'fecha_factura': registro.fecha_factura.strftime('%Y-%m-%d') if registro.fecha_factura else None,
        'precio': float(registro.precio) if registro.precio is not None else 0,
        'forma_pago': registro.forma_pago,
        'servicio': registro.servicio,
        'nro_identificacion': registro.nro_identificacion,
        'nombre_paciente': registro.nombre_paciente,
        'acuerdo_comercial': registro.acuerdo_comercial,
        'empresa_mision': registro.empresa_mision,
        'sede': registro.sede,
        'nombre_vendedor': registro.nombre_vendedor,
        'fecha_creacion_orden': registro.fecha_creacion_orden.strftime('%Y-%m-%d %H:%M') if registro.fecha_creacion_orden else None,
        'usuario_creacion': registro.usuario_creacion,
        'estado_orden': registro.estado_orden,
        'estado_gestion': registro.estado_gestion,
        'es_editable': (registro.estado_gestion or '').upper() == 'CARGADA',
        'fecha_anulacion': registro.fecha_anulacion.strftime('%Y-%m-%d') if registro.fecha_anulacion else None,
        'archivo_origen': registro.archivo_origen,
        'created_at': registro.created_at.strftime('%Y-%m-%d %H:%M:%S') if registro.created_at else None,
    }


def _orden_num_expr_atenciones():
    return cast(
        func.nullif(
            func.regexp_replace(
                func.coalesce(AtencionDiaDetalle.nro_orden, ''),
                r'[^0-9]',
                '',
                'g',
            ),
            '',
        ),
        BigInteger,
    )


def _claves_grupo_orden_atenciones():
    return {
        'cliente_group': func.coalesce(AtencionDiaDetalle.cliente_id, 0),
        'vendedor_group': func.coalesce(AtencionDiaDetalle.vendedor_id, 0),
        'orden_group': func.coalesce(
            func.nullif(AtencionDiaDetalle.nro_orden, ''),
            func.concat('__sin_orden__', cast(AtencionDiaDetalle.id, String)),
        ),
    }


def _coincide_detalle_prefactura_con_atencion(detalle: PrefacturaComercialDetalle, registro: AtencionDiaDetalle) -> bool:
    if detalle is None or registro is None:
        return False
    if detalle.prefactura is None or detalle.prefactura.cliente_id != registro.cliente_id:
        return False
    if not detalle.fecha_programada or not registro.fecha_creacion_orden:
        return False
    if detalle.fecha_programada.date() != registro.fecha_creacion_orden.date():
        return False
    if _normalizar_match(detalle.paciente_documento) != _normalizar_match(registro.nro_identificacion):
        return False
    if _normalizar_match(detalle.nombre_item) != _normalizar_match(registro.servicio):
        return False
    return True


def _sincronizar_cruce_prefacturas_con_atencion(registro: AtencionDiaDetalle) -> None:
    if registro is None or registro.id is None or registro.cliente_id is None:
        return
    if not _detalle_prefactura_manual_disponible():
        return

    try:
        candidatos = (
            PrefacturaComercialDetalle.query
            .join(PrefacturaComercial, PrefacturaComercial.id == PrefacturaComercialDetalle.prefactura_id)
            .filter(
                PrefacturaComercial.origen == PREF_ORIGEN_MANUAL,
                PrefacturaComercial.cliente_id == registro.cliente_id,
                PrefacturaComercialDetalle.atencion_dia_id.is_(None),
            )
            .order_by(
                PrefacturaComercialDetalle.fecha_programada.asc(),
                PrefacturaComercialDetalle.id.asc(),
            )
            .all()
        )
    except (ProgrammingError, OperationalError):
        return

    for detalle in candidatos:
        if not _coincide_detalle_prefactura_con_atencion(detalle, registro):
            continue
        detalle.atencion_dia_id = registro.id
        detalle.estado_cruce = PREF_DETALLE_CRUCE_CRUZADO
        detalle.cruzado_at = datetime.utcnow()
        break


def _desvincular_prefacturas_de_atencion(registro: AtencionDiaDetalle) -> None:
    if registro is None or registro.id is None:
        return
    if not _detalle_prefactura_manual_disponible():
        return

    try:
        detalles = PrefacturaComercialDetalle.query.filter_by(atencion_dia_id=registro.id).all()
    except (ProgrammingError, OperationalError):
        return
    for detalle in detalles:
        detalle.atencion_dia_id = None
        detalle.estado_cruce = PREF_DETALLE_CRUCE_PENDIENTE
        detalle.cruzado_at = None


def _intentar_cruzar_prefactura_manual(prefactura: PrefacturaComercial | None) -> None:
    if prefactura is None or (prefactura.origen or PREF_ORIGEN_ATENCIONES).upper() != PREF_ORIGEN_MANUAL:
        return
    if not _detalle_prefactura_manual_disponible():
        return

    try:
        detalles = (
            prefactura.detalles
            .filter(PrefacturaComercialDetalle.atencion_dia_id.is_(None))
            .order_by(PrefacturaComercialDetalle.id.asc())
            .all()
        )
    except (ProgrammingError, OperationalError):
        return
    if not detalles:
        return

    fecha_programada = prefactura.fecha_programada or prefactura.fecha_desde
    if fecha_programada is None:
        return

    registros = (
        AtencionDiaDetalle.query
        .filter(
            AtencionDiaDetalle.cliente_id == prefactura.cliente_id,
            AtencionDiaDetalle.fecha_creacion_orden >= fecha_programada.replace(hour=0, minute=0, second=0, microsecond=0),
            AtencionDiaDetalle.fecha_creacion_orden <= fecha_programada.replace(hour=23, minute=59, second=59, microsecond=999999),
        )
        .order_by(AtencionDiaDetalle.id.asc())
        .all()
    )

    usados = {detalle.atencion_dia_id for detalle in prefactura.detalles if detalle.atencion_dia_id}
    for detalle in detalles:
        for registro in registros:
            if registro.id in usados:
                continue
            if not _coincide_detalle_prefactura_con_atencion(detalle, registro):
                continue
            detalle.atencion_dia_id = registro.id
            detalle.estado_cruce = PREF_DETALLE_CRUCE_CRUZADO
            detalle.cruzado_at = datetime.utcnow()
            usados.add(registro.id)
            break


def _parse_fecha_iso(valor, field_name):
    texto = _normalizar(valor)
    if not texto:
        return None
    try:
        return datetime.strptime(texto, '%Y-%m-%d')
    except ValueError as exc:
        raise ValueError(f'Formato de {field_name} invalido (YYYY-MM-DD)') from exc


def _validar_periodo_cargue(desde, hasta):
    if bool(desde) != bool(hasta):
        raise ValueError('Debes registrar tanto la fecha inicial como la fecha final del periodo')
    if desde and hasta and hasta < desde:
        raise ValueError('La fecha final del periodo no puede ser anterior a la fecha inicial')


def _formatear_periodo_cargue(desde, hasta):
    if not desde or not hasta:
        return None
    return f'{desde.strftime("%Y-%m-%d")} a {hasta.strftime("%Y-%m-%d")}'


def _serializar_periodo_cargue(desde, hasta, *, source='CARGUE'):
    if not desde or not hasta:
        return None
    return {
        'fecha_desde': desde.strftime('%Y-%m-%d'),
        'fecha_hasta': hasta.strftime('%Y-%m-%d'),
        'label': _formatear_periodo_cargue(desde, hasta),
        'source': source,
        'key': f'{desde.strftime("%Y-%m-%d")}|{hasta.strftime("%Y-%m-%d")}|{source}',
    }


def _parse_int_optional(valor, field_name):
    if valor in (None, ''):
        return None
    try:
        return int(valor)
    except (TypeError, ValueError) as exc:
        raise ValueError(f'El campo {field_name} debe ser numerico') from exc


def _resolver_cliente_desde_payload(data, vendedor_scope):
    cliente_id = _parse_int_optional(data.get('cliente_id'), 'cliente_id')
    if cliente_id is not None:
        cliente = _clientes_visibles_query(vendedor_scope).filter(ClienteComercial.id == cliente_id).first()
        if cliente is None:
            raise ValueError('El cliente seleccionado no esta disponible para tu perfil')
        return cliente

    clientes_lookup = _construir_lookup_clientes_visibles(vendedor_scope)
    for campo in ('acuerdo_comercial', 'empresa_mision'):
        normalized = _normalizar_match(data.get(campo))
        if not normalized or normalized in _VALORES_EMPRESA_GENERICOS:
            continue
        cliente = clientes_lookup.get(normalized)
        if cliente is not None:
            return cliente
    return None


def _resolver_vendedor_desde_payload(data, vendedor_scope, cliente):
    if not _is_admin_user():
        if vendedor_scope is None:
            raise PermissionError('No tienes un vendedor asociado para registrar atenciones')
        return vendedor_scope

    vendedor_id = _parse_int_optional(data.get('vendedor_id'), 'vendedor_id')
    if vendedor_id is not None:
        vendedor = Vendedor.query.get(vendedor_id)
        if vendedor is None:
            raise ValueError('El vendedor seleccionado no existe')
        return vendedor

    nombre_vendedor = _normalizar(data.get('nombre_vendedor'))
    if nombre_vendedor:
        vendedor = _construir_lookup_vendedores().get(_normalizar_match(nombre_vendedor))
        if vendedor is not None:
            return vendedor

    if cliente is not None and cliente.vendedor is not None:
        return cliente.vendedor
    return None


def _build_atencion_dia_payload(data, vendedor_scope):
    fecha_creacion_orden = _parse_fecha_iso(data.get('fecha_creacion_orden'), 'fecha_creacion_orden') or datetime.utcnow()
    fecha_factura = _parse_fecha_iso(data.get('fecha_factura'), 'fecha_factura')
    precio = _parse_precio(data.get('precio'))
    if precio is None:
        raise ValueError('El precio es obligatorio')
    if precio < 0:
        raise ValueError('El precio no puede ser negativo')

    nro_identificacion = _normalizar(data.get('nro_identificacion'))
    nombre_paciente = _normalizar(data.get('nombre_paciente'))
    servicio = _normalizar(data.get('servicio'))
    if not nro_identificacion:
        raise ValueError('El numero de identificacion es obligatorio')
    if not nombre_paciente:
        raise ValueError('El nombre del paciente es obligatorio')
    if not servicio:
        raise ValueError('El servicio es obligatorio')

    cliente = _resolver_cliente_desde_payload(data, vendedor_scope)
    vendedor = _resolver_vendedor_desde_payload(data, vendedor_scope, cliente)
    estado_orden = (_normalizar(data.get('estado_orden')) or 'PROCESADA').upper()
    forma_pago = (_normalizar(data.get('forma_pago')) or None)
    if forma_pago:
        forma_pago = forma_pago.upper()

    acuerdo_comercial = _normalizar(data.get('acuerdo_comercial'))
    empresa_mision = _normalizar(data.get('empresa_mision'))
    nombre_vendedor = _normalizar(data.get('nombre_vendedor'))
    if vendedor is not None and not nombre_vendedor:
        nombre_vendedor = vendedor.nombre

    fecha_anulacion = None
    if estado_orden == 'ANULADA':
        fecha_anulacion = _parse_fecha_iso(data.get('fecha_anulacion'), 'fecha_anulacion') or fecha_creacion_orden

    return {
        'cliente_id': cliente.id if cliente is not None else None,
        'vendedor_id': vendedor.id if vendedor is not None else None,
        'nro_orden': _normalizar(data.get('nro_orden')),
        'nro_factura': _normalizar(data.get('nro_factura')),
        'fecha_factura': fecha_factura,
        'precio': precio,
        'forma_pago': forma_pago,
        'servicio': servicio,
        'nro_identificacion': nro_identificacion,
        'nombre_paciente': nombre_paciente,
        'acuerdo_comercial': acuerdo_comercial,
        'empresa_mision': empresa_mision,
        'sede': _normalizar(data.get('sede')),
        'nombre_vendedor': nombre_vendedor,
        'fecha_creacion_orden': fecha_creacion_orden,
        'usuario_creacion': _normalizar(data.get('usuario_creacion')) or getattr(current_user, 'usuario', None),
        'estado_orden': estado_orden,
        'estado_gestion': 'CARGADA',
        'fecha_anulacion': fecha_anulacion,
        'archivo_origen': _normalizar(data.get('archivo_origen')) or NOMBRE_CARGUE_MANUAL,
    }


@comercial_bp.route('/cargue-atenciones', methods=['POST'])
@login_required
def cargar_atenciones_dia():
    """Recibe un archivo Excel y carga las atenciones del dÃÆÃÂ­a."""
    try:
        _require_commercial_permission(PERMISO_CARGUE_ATENCIONES)
    except PermissionError as exc:
        return jsonify({'error': str(exc)}), 403

    if 'archivo' not in request.files:
        return jsonify({'error': 'No se enviÃÆÃÂ³ ningÃÆÃÂºn archivo'}), 400

    archivo = request.files['archivo']
    nombre = archivo.filename or 'sin_nombre.xlsx'
    if not nombre.lower().endswith('.xlsx'):
        return jsonify({'error': 'Solo se aceptan archivos .xlsx'}), 400

    try:
        periodo_desde = _parse_fecha_iso(request.form.get('periodo_desde'), 'periodo_desde')
        periodo_hasta = _parse_fecha_iso(request.form.get('periodo_hasta'), 'periodo_hasta')
        _validar_periodo_cargue(periodo_desde, periodo_hasta)
        if not periodo_desde or not periodo_hasta:
            return jsonify({'error': 'Debes indicar el periodo inicial del cargue (fecha desde y fecha hasta)'}), 400
    except ValueError as exc:
        return jsonify({'error': str(exc)}), 400

    try:
        contenido = archivo.read()
        filas = _leer_excel_atenciones(contenido, nombre)
        registros = _extraer_registros_excel(filas)
    except ValueError as exc:
        return jsonify({'error': str(exc)}), 400
    except Exception as exc:
        logger.error('Error inesperado leyendo Excel de atenciones: %s', exc)
        return jsonify({'error': f'No se pudo leer el archivo: {exc}'}), 400

    clientes_lookup = _construir_lookup_clientes()
    vendedores_lookup = _construir_lookup_vendedores()
    existentes = set(
        db.session.query(
            AtencionDiaDetalle.nro_orden,
            AtencionDiaDetalle.servicio,
            AtencionDiaDetalle.nro_identificacion,
        ).all()
    )

    try:
        cargue, legacy_cargue = _crear_cargue_atenciones(
            nombre,
            len(registros),
            current_user.id,
            periodo_desde,
            periodo_hasta,
        )
    except (ProgrammingError, OperationalError) as exc:
        db.session.rollback()
        return jsonify({'error': f'No se pudo iniciar el cargue en base de datos: {exc}'}), 500

    importadas = 0
    duplicadas = 0
    errores = 0
    relacionadas_cliente = 0
    relacionadas_vendedor = 0
    registros_nuevos: list[AtencionDiaDetalle] = []

    for registro in registros:
        try:
            nro_orden = _normalizar(registro.get('NÃâÃÂ°. Orden Servicio'))
            servicio = _normalizar(registro.get('Nombre del Producto o Servicio'))
            nro_identificacion = _normalizar(registro.get('NÃâÃÂ°. de IdentificaciÃÆÃÂ³n'))

            clave = (nro_orden, servicio, nro_identificacion)
            if clave in existentes:
                duplicadas += 1
                continue

            cliente = _cliente_desde_registro(registro, clientes_lookup)
            vendedor = cliente.vendedor if cliente and cliente.vendedor is not None else _vendedor_desde_registro(registro, vendedores_lookup)

            detalle = AtencionDiaDetalle(
                cargue_id=cargue.id,
                cliente_id=cliente.id if cliente else None,
                vendedor_id=vendedor.id if vendedor else None,
                nro_orden=nro_orden,
                nro_factura=_normalizar(registro.get('NÃâÃÂ°. Factura')),
                fecha_factura=_parse_fecha(registro.get('Fecha de Factura')),
                precio=_parse_precio(registro.get('Precio')),
                forma_pago=_normalizar(registro.get('FormaPago')),
                servicio=servicio,
                nro_identificacion=nro_identificacion,
                nombre_paciente=_normalizar(registro.get('Nombre del Paciente')),
                acuerdo_comercial=_normalizar(registro.get('Nombre del Acuerdo Comercial')),
                empresa_mision=_normalizar(registro.get('Empresa en MisiÃÆÃÂ³n')),
                sede=_normalizar(registro.get('Sede')),
                nombre_vendedor=_normalizar(registro.get('Nombre del Vendedor')),
                fecha_creacion_orden=_parse_fecha(registro.get('Fecha de CreaciÃÆÃÂ³n Orden Servicio')),
                usuario_creacion=_normalizar(registro.get('Usuario de CreaciÃÆÃÂ³n Orden Servicio')),
                estado_orden=_normalizar(registro.get('Estado de la Orden Servicio')),
                estado_gestion='CARGADA',
                fecha_anulacion=_parse_fecha(registro.get('Fecha de AnulaciÃÆÃÂ³n Orden Servicio')),
                archivo_origen=nombre,
            )
            db.session.add(detalle)
            existentes.add(clave)
            registros_nuevos.append(detalle)
            importadas += 1
            if cliente is not None:
                relacionadas_cliente += 1
            if vendedor is not None:
                relacionadas_vendedor += 1
        except Exception as exc:
            logger.warning('Error procesando fila de atenciones: %s', exc)
            errores += 1

    try:
        _actualizar_resumen_cargue(cargue, importadas, duplicadas, errores, legacy_mode=legacy_cargue)
        db.session.flush()
        for detalle in registros_nuevos:
            _sincronizar_cruce_prefacturas_con_atencion(detalle)
        db.session.commit()
    except (ProgrammingError, OperationalError) as exc:
        db.session.rollback()
        mensaje_estado = _mensaje_error_estado_gestion_atenciones(exc)
        if mensaje_estado:
            return jsonify({'error': mensaje_estado}), 500
        return jsonify({'error': f'No se pudo guardar el cargue en base de datos: {exc}'}), 500
    except Exception as exc:
        db.session.rollback()
        logger.error('Error finalizando cargue de atenciones: %s', exc)
        return jsonify({'error': 'No se pudo completar el cargue de atenciones'}), 500

    logger.info(
        'Cargue atenciones: archivo=%s importadas=%d duplicadas=%d errores=%d relacionadas_cliente=%d relacionadas_vendedor=%d',
        nombre,
        importadas,
        duplicadas,
        errores,
        relacionadas_cliente,
        relacionadas_vendedor,
    )

    return jsonify({
        'mensaje': 'Cargue completado',
        'nombre_archivo': nombre,
        'total_filas': cargue.total_filas,
        'importadas': importadas,
        'duplicadas': duplicadas,
        'errores': errores,
        'relacionadas_cliente': relacionadas_cliente,
        'sin_cliente': max(importadas - relacionadas_cliente, 0),
        'relacionadas_vendedor': relacionadas_vendedor,
        'sin_vendedor': max(importadas - relacionadas_vendedor, 0),
        'periodo': _serializar_periodo_cargue(periodo_desde, periodo_hasta),
    }), 201


@comercial_bp.route('/cargue-atenciones/historial', methods=['GET'])
@login_required
def historial_cargues_atenciones():
    """ÃÆÃÂ¡ltimos 20 cargues realizados."""
    try:
        _require_commercial_permission(PERMISO_CONSULTA_ATENCIONES)
    except PermissionError as exc:
        return jsonify({'error': str(exc)}), 403

    vendedor_scope = _resolver_vendedor_usuario_actual()
    scope = _scope_descriptor(vendedor_scope)

    if not _is_admin_user() and vendedor_scope is None:
        return jsonify({'scope': scope, 'registros': []}), 200

    if not _columnas_periodo_cargue_disponibles():
        if _is_admin_user():
            rows = db.session.execute(text("""
                select c.id, c.nombre_archivo, c.total_filas, c.filas_importadas,
                       c.filas_duplicadas, c.filas_error, c.created_at, u.usuario
                from cargue_atenciones_dia c
                left join usuarios u on u.id = c.usuario_id
                order by c.created_at desc
                limit 20
            """)).fetchall()
            registros = [{
                'id': row.id,
                'nombre_archivo': row.nombre_archivo,
                'total_filas': row.total_filas,
                'importadas': row.filas_importadas,
                'duplicadas': row.filas_duplicadas,
                'errores': row.filas_error,
                'usuario': row.usuario or 'Sistema',
                'fecha': row.created_at.strftime('%Y-%m-%d %H:%M') if row.created_at else None,
                'periodo': None,
            } for row in rows]
            return jsonify({'scope': scope, 'registros': registros}), 200

        rows = db.session.execute(text("""
            select c.id, c.nombre_archivo, c.created_at, u.usuario,
                   count(d.id) as visibles
            from cargue_atenciones_dia c
            join atenciones_dia_detalle d on d.cargue_id = c.id
            left join usuarios u on u.id = c.usuario_id
            where d.vendedor_id = :vendedor_id
            group by c.id, c.nombre_archivo, c.created_at, u.usuario
            order by c.created_at desc
            limit 20
        """), {'vendedor_id': vendedor_scope.id}).fetchall()
        registros = [{
            'id': row.id,
            'nombre_archivo': row.nombre_archivo,
            'total_filas': int(row.visibles or 0),
            'importadas': int(row.visibles or 0),
            'duplicadas': None,
            'errores': None,
            'usuario': row.usuario or 'Sistema',
            'fecha': row.created_at.strftime('%Y-%m-%d %H:%M') if row.created_at else None,
            'periodo': None,
        } for row in rows]
        return jsonify({'scope': scope, 'registros': registros}), 200

    if _is_admin_user():
        cargues = CargueAtencionDia.query.order_by(CargueAtencionDia.created_at.desc()).limit(20).all()
        registros = [{
            'id': cargue.id,
            'nombre_archivo': cargue.nombre_archivo,
            'total_filas': cargue.total_filas,
            'importadas': cargue.filas_importadas,
            'duplicadas': cargue.filas_duplicadas,
            'errores': cargue.filas_error,
            'usuario': cargue.usuario.usuario if cargue.usuario else 'Sistema',
            'fecha': cargue.created_at.strftime('%Y-%m-%d %H:%M') if cargue.created_at else None,
            'periodo': _formatear_periodo_cargue(cargue.periodo_desde, cargue.periodo_hasta),
        } for cargue in cargues]
        return jsonify({'scope': scope, 'registros': registros}), 200

    cargues = (
        db.session.query(CargueAtencionDia)
        .join(AtencionDiaDetalle, AtencionDiaDetalle.cargue_id == CargueAtencionDia.id)
        .filter(AtencionDiaDetalle.vendedor_id == vendedor_scope.id)
        .order_by(CargueAtencionDia.created_at.desc())
        .distinct()
        .limit(20)
        .all()
    )

    registros = []
    for cargue in cargues:
        visibles = AtencionDiaDetalle.query.filter_by(cargue_id=cargue.id, vendedor_id=vendedor_scope.id).count()
        registros.append({
            'id': cargue.id,
            'nombre_archivo': cargue.nombre_archivo,
            'total_filas': visibles,
            'importadas': visibles,
            'duplicadas': None,
            'errores': None,
            'usuario': cargue.usuario.usuario if cargue.usuario else 'Sistema',
            'fecha': cargue.created_at.strftime('%Y-%m-%d %H:%M') if cargue.created_at else None,
            'periodo': _formatear_periodo_cargue(cargue.periodo_desde, cargue.periodo_hasta),
        })

    return jsonify({'scope': scope, 'registros': registros}), 200


@comercial_bp.route('/cargue-atenciones/periodos', methods=['GET'])
@login_required
def listar_periodos_cargue_atenciones():
    try:
        _require_commercial_permission(PERMISO_CONSULTA_ATENCIONES)
    except PermissionError as exc:
        return jsonify({'error': str(exc)}), 403

    vendedor_scope = _resolver_vendedor_usuario_actual()
    scope = _scope_descriptor(vendedor_scope)
    if not _is_admin_user() and vendedor_scope is None:
        return jsonify({'scope': scope, 'periodos': []}), 200

    periodos = []
    vistos = set()

    if _columnas_periodo_cargue_disponibles():
        query_cargues = CargueAtencionDia.query.filter(
            CargueAtencionDia.periodo_desde.isnot(None),
            CargueAtencionDia.periodo_hasta.isnot(None),
        )
        if not _is_admin_user():
            query_cargues = (
                query_cargues
                .join(AtencionDiaDetalle, AtencionDiaDetalle.cargue_id == CargueAtencionDia.id)
                .filter(AtencionDiaDetalle.vendedor_id == vendedor_scope.id)
            )

        for cargue in query_cargues.order_by(CargueAtencionDia.periodo_desde.desc(), CargueAtencionDia.periodo_hasta.desc()).all():
            periodo = _serializar_periodo_cargue(cargue.periodo_desde, cargue.periodo_hasta, source='CARGUE')
            if not periodo:
                continue
            key = (periodo['fecha_desde'], periodo['fecha_hasta'])
            if key in vistos:
                continue
            vistos.add(key)
            periodos.append(periodo)

    if _columnas_prefactura_avanzadas_disponibles():
        query_prefacturas = PrefacturaComercial.query.filter(
            PrefacturaComercial.fecha_desde.isnot(None),
            PrefacturaComercial.fecha_hasta.isnot(None),
        )
        if not _is_admin_user():
            cliente_ids = [c.id for c in ClienteComercial.query.filter_by(vendedor_id=vendedor_scope.id).all()]
            if not cliente_ids:
                return jsonify({'scope': scope, 'periodos': periodos}), 200
            query_prefacturas = query_prefacturas.filter(PrefacturaComercial.cliente_id.in_(cliente_ids))

        pref_rows = [
            {
                'fecha_desde': pref.fecha_desde,
                'fecha_hasta': pref.fecha_hasta,
            }
            for pref in query_prefacturas.order_by(PrefacturaComercial.fecha_desde.desc(), PrefacturaComercial.fecha_hasta.desc()).all()
        ]
    else:
        pref_rows = [
            {
                'fecha_desde': row.fecha_desde,
                'fecha_hasta': row.fecha_hasta,
                'cliente_id': row.cliente_id,
            }
            for row in db.session.execute(text("""
                select cliente_id, fecha_desde, fecha_hasta
                from prefacturas_comerciales
                where fecha_desde is not null and fecha_hasta is not null
                order by fecha_desde desc, fecha_hasta desc
            """)).fetchall()
        ]
        if not _is_admin_user():
            cliente_ids = {c.id for c in ClienteComercial.query.filter_by(vendedor_id=vendedor_scope.id).all()}
            if not cliente_ids:
                return jsonify({'scope': scope, 'periodos': periodos}), 200
            pref_rows = [row for row in pref_rows if row.get('cliente_id') in cliente_ids]

    for pref in pref_rows:
        periodo = _serializar_periodo_cargue(pref['fecha_desde'], pref['fecha_hasta'], source='PREFACTURA')
        if not periodo:
            continue
        key = (periodo['fecha_desde'], periodo['fecha_hasta'])
        if key in vistos:
            continue
        vistos.add(key)
        periodos.append(periodo)

    periodos.sort(key=lambda item: (item['fecha_desde'], item['fecha_hasta']), reverse=True)
    return jsonify({'scope': scope, 'periodos': periodos}), 200


@comercial_bp.route('/cargue-atenciones/clientes-sugeridos', methods=['GET'])
@login_required
def clientes_sugeridos_atenciones():
    """Devuelve clientes visibles para autocompletar el filtro Cliente / Acuerdo."""
    try:
        _require_commercial_permission(PERMISO_CONSULTA_ATENCIONES)
    except PermissionError as exc:
        return jsonify({'error': str(exc)}), 403

    vendedor_scope = _resolver_vendedor_usuario_actual()
    scope = _scope_descriptor(vendedor_scope)
    if not _is_admin_user() and vendedor_scope is None:
        return jsonify({'scope': scope, 'clientes': []}), 200

    query_text = request.args.get('q', '').strip()
    if not query_text:
        return jsonify({'scope': scope, 'clientes': []}), 200

    query = _clientes_visibles_query(vendedor_scope)
    query = query.filter(
        or_(
            ClienteComercial.razon_social.ilike(f'%{query_text}%'),
            ClienteComercial.nombre_comercial.ilike(f'%{query_text}%'),
            ClienteComercial.nit.ilike(f'%{query_text}%'),
        )
    )

    clientes = query.order_by(
        ClienteComercial.activo.desc(),
        ClienteComercial.razon_social.asc(),
    ).limit(15).all()

    return jsonify({
        'scope': scope,
        'clientes': [
            {
                'id': cliente.id,
                'razon_social': cliente.razon_social,
                'nombre_comercial': cliente.nombre_comercial,
                'nit': cliente.nit,
                'vendedor_id': cliente.vendedor_id,
                'vendedor_nombre': cliente.vendedor.nombre if cliente.vendedor else None,
            }
            for cliente in clientes
        ],
    }), 200


@comercial_bp.route('/cargue-atenciones/consulta', methods=['GET'])
@login_required
def consultar_atenciones_dia():
    """Consulta de atenciones cargadas con alcance automÃÆÃÂ¡tico por vendedor."""
    try:
        _require_commercial_permission(PERMISO_CONSULTA_ATENCIONES)
    except PermissionError as exc:
        return jsonify({'error': str(exc)}), 403

    acuerdo = request.args.get('acuerdo', '').strip()
    vendedor = request.args.get('vendedor', '').strip()
    condicion_comercial = request.args.get('condicion_comercial', '').strip()
    estado = request.args.get('estado', '').strip()
    fecha_desde = request.args.get('fecha_desde', '').strip()
    fecha_hasta = request.args.get('fecha_hasta', '').strip()
    cargue_id = request.args.get('cargue_id', type=int)
    cliente_id = request.args.get('cliente_id', type=int)
    page = max(1, int(request.args.get('page', 1)))
    per_page = min(200, max(10, int(request.args.get('per_page', 50))))

    vendedor_scope = _resolver_vendedor_usuario_actual()
    scope = _scope_descriptor(vendedor_scope)
    if not _is_admin_user() and vendedor_scope is None:
        return jsonify({
            'scope': scope,
            'search_required': False,
            'total': 0,
            'page': page,
            'per_page': per_page,
            'pages': 0,
            'registros': [],
        }), 200

    has_filters = any([
        acuerdo,
        vendedor,
        condicion_comercial,
        estado,
        fecha_desde,
        fecha_hasta,
        cargue_id,
        cliente_id,
    ])
    if not has_filters:
        return jsonify({
            'scope': scope,
            'search_required': True,
            'vendedor_scope_id': vendedor_scope.id if vendedor_scope else None,
            'total': 0,
            'page': 1,
            'per_page': per_page,
            'pages': 0,
            'registros': [],
        }), 200

    query = AtencionDiaDetalle.query.outerjoin(AtencionDiaDetalle.cliente).outerjoin(AtencionDiaDetalle.vendedor)

    if not _is_admin_user():
        query = query.filter(AtencionDiaDetalle.vendedor_id == vendedor_scope.id)

    if acuerdo:
        query = query.filter(
            or_(
                AtencionDiaDetalle.acuerdo_comercial.ilike(f'%{acuerdo}%'),
                AtencionDiaDetalle.empresa_mision.ilike(f'%{acuerdo}%'),
                ClienteComercial.razon_social.ilike(f'%{acuerdo}%'),
                ClienteComercial.nombre_comercial.ilike(f'%{acuerdo}%'),
                ClienteComercial.nit.ilike(f'%{acuerdo}%'),
            )
        )
    if vendedor:
        query = query.filter(
            or_(
                AtencionDiaDetalle.nombre_vendedor.ilike(f'%{vendedor}%'),
                Vendedor.nombre.ilike(f'%{vendedor}%'),
            )
        )
    if condicion_comercial:
        query = query.filter(ClienteComercial.condicion_comercial == condicion_comercial.upper())
    if estado:
        query = query.filter(AtencionDiaDetalle.estado_gestion == estado.upper())
    if cargue_id:
        query = query.filter(AtencionDiaDetalle.cargue_id == cargue_id)
    if cliente_id:
        query = query.filter(AtencionDiaDetalle.cliente_id == cliente_id)
    if fecha_desde:
        try:
            query = query.filter(AtencionDiaDetalle.fecha_creacion_orden >= datetime.strptime(fecha_desde, '%Y-%m-%d'))
        except ValueError:
            pass
    if fecha_hasta:
        try:
            fecha_fin = datetime.strptime(f'{fecha_hasta} 23:59:59', '%Y-%m-%d %H:%M:%S')
            query = query.filter(AtencionDiaDetalle.fecha_creacion_orden <= fecha_fin)
        except ValueError:
            pass

    try:
        total_registros = query.count()
        orden_num_expr = _orden_num_expr_atenciones()
        group_exprs = _claves_grupo_orden_atenciones()
        fecha_orden_expr = func.min(AtencionDiaDetalle.fecha_creacion_orden)

        grupos_subquery = (
            query.with_entities(
                group_exprs['cliente_group'].label('cliente_group'),
                group_exprs['vendedor_group'].label('vendedor_group'),
                group_exprs['orden_group'].label('orden_group'),
                orden_num_expr.label('orden_num'),
                fecha_orden_expr.label('fecha_orden'),
            )
            .group_by(
                group_exprs['cliente_group'],
                group_exprs['vendedor_group'],
                group_exprs['orden_group'],
                orden_num_expr,
            )
            .subquery()
        )

        total_ordenes = db.session.query(func.count()).select_from(grupos_subquery).scalar() or 0
        grupos_paginados = (
            db.session.query(grupos_subquery)
            .order_by(
                grupos_subquery.c.orden_num.asc().nullslast(),
                grupos_subquery.c.orden_group.asc().nullslast(),
                grupos_subquery.c.fecha_orden.asc().nullslast(),
                grupos_subquery.c.cliente_group.asc(),
                grupos_subquery.c.vendedor_group.asc(),
            )
            .offset((page - 1) * per_page)
            .limit(per_page)
            .subquery()
        )

        registros = (
            query.join(
                grupos_paginados,
                and_(
                    group_exprs['cliente_group'] == grupos_paginados.c.cliente_group,
                    group_exprs['vendedor_group'] == grupos_paginados.c.vendedor_group,
                    group_exprs['orden_group'] == grupos_paginados.c.orden_group,
                ),
            )
            .order_by(
                grupos_paginados.c.orden_num.asc().nullslast(),
                grupos_paginados.c.orden_group.asc().nullslast(),
                grupos_paginados.c.fecha_orden.asc().nullslast(),
                AtencionDiaDetalle.nro_identificacion.asc().nullslast(),
                AtencionDiaDetalle.nombre_paciente.asc().nullslast(),
                AtencionDiaDetalle.servicio.asc().nullslast(),
                AtencionDiaDetalle.id.asc(),
            )
            .all()
        )
    except (ProgrammingError, OperationalError) as exc:
        db.session.rollback()
        mensaje_estado = _mensaje_error_estado_gestion_atenciones(exc)
        logger.exception('Error consultando atenciones dia')
        return jsonify({'error': mensaje_estado or 'No se pudo consultar la informacion cargada'}), 500

    return jsonify({
        'scope': scope,
        'search_required': False,
        'vendedor_scope_id': vendedor_scope.id if vendedor_scope else None,
        'total': total_ordenes,
        'total_ordenes': total_ordenes,
        'total_registros': total_registros,
        'page': page,
        'per_page': per_page,
        'pages': (total_ordenes + per_page - 1) // per_page,
        'registros': [_serialize_atencion_dia(registro) for registro in registros],
    }), 200


# ---------------------------------------------------------------------------
# GET /api/comercial/atenciones-dia/<id>  — detalle de un registro
# ---------------------------------------------------------------------------
@comercial_bp.route('/atenciones-dia/<int:reg_id>', methods=['GET'])
@login_required
def obtener_atencion_dia(reg_id):
    try:
        _require_commercial_permission(PERMISO_CONSULTA_ATENCIONES)
    except PermissionError as exc:
        return jsonify({'error': str(exc)}), 403

    vendedor_scope = _resolver_vendedor_usuario_actual()
    if not _is_admin_user() and vendedor_scope is None:
        return jsonify({'error': 'No tienes un vendedor asociado'}), 403

    try:
        reg = AtencionDiaDetalle.query.get_or_404(reg_id)
    except (ProgrammingError, OperationalError) as exc:
        db.session.rollback()
        mensaje_estado = _mensaje_error_estado_gestion_atenciones(exc)
        logger.exception('Error obteniendo detalle de atencion dia %s', reg_id)
        return jsonify({'error': mensaje_estado or 'No se pudo consultar la atencion'}), 500

    try:
        _asegurar_acceso_registro_atencion(reg, vendedor_scope)
    except PermissionError as exc:
        return jsonify({'error': str(exc)}), 403

    return jsonify({'registro': _serialize_atencion_dia(reg)}), 200


# ---------------------------------------------------------------------------
# POST /api/comercial/atenciones-dia  — crear registro manual
# ---------------------------------------------------------------------------
@comercial_bp.route('/atenciones-dia', methods=['POST'])
@login_required
def crear_atencion_dia_manual():
    try:
        _require_commercial_permission(PERMISO_CARGUE_ATENCIONES)
        vendedor_scope = _resolver_vendedor_usuario_actual()
        if not _is_admin_user() and vendedor_scope is None:
            return jsonify({'error': 'No tienes un vendedor asociado'}), 403

        data = request.get_json() or {}
        payload = _build_atencion_dia_payload(data, vendedor_scope)
        payload['estado_gestion'] = 'CARGADA'

        cargue = CargueAtencionDia(
            nombre_archivo=NOMBRE_CARGUE_MANUAL,
            total_filas=1,
            filas_importadas=1,
            filas_duplicadas=0,
            filas_error=0,
            usuario_id=getattr(current_user, 'id', None),
        )
        db.session.add(cargue)
        db.session.flush()

        registro = AtencionDiaDetalle(
            cargue_id=cargue.id,
            **payload,
        )
        db.session.add(registro)
        db.session.flush()
        _sincronizar_cruce_prefacturas_con_atencion(registro)
        db.session.commit()
    except ValueError as exc:
        db.session.rollback()
        return jsonify({'error': str(exc)}), 400
    except PermissionError as exc:
        db.session.rollback()
        return jsonify({'error': str(exc)}), 403
    except Exception as exc:
        db.session.rollback()
        logger.error('Error creando atencion manual: %s', exc)
        return jsonify({'error': 'No se pudo crear la atencion'}), 500

    return jsonify({'mensaje': 'Atencion creada', 'registro': _serialize_atencion_dia(registro)}), 201


@comercial_bp.route('/atenciones-dia/<int:reg_id>', methods=['PUT'])
@login_required
def editar_atencion_dia(reg_id):
    try:
        _require_commercial_permission(PERMISO_EDICION_ATENCIONES)
    except PermissionError as exc:
        return jsonify({'error': str(exc)}), 403

    vendedor_scope = _resolver_vendedor_usuario_actual()
    if not _is_admin_user() and vendedor_scope is None:
        return jsonify({'error': 'No tienes un vendedor asociado'}), 403

    reg = AtencionDiaDetalle.query.get_or_404(reg_id)
    try:
        _asegurar_acceso_registro_atencion(reg, vendedor_scope)
        if (reg.estado_gestion or '').upper() != 'CARGADA':
            return jsonify({'error': 'Solo se pueden editar atenciones en estado CARGADA'}), 409
        data = request.get_json() or {}
        payload = _build_atencion_dia_payload(data, vendedor_scope)
        payload['archivo_origen'] = reg.archivo_origen or payload.get('archivo_origen')
        payload['estado_gestion'] = reg.estado_gestion or 'CARGADA'
    except ValueError as exc:
        return jsonify({'error': str(exc)}), 400
    except PermissionError as exc:
        return jsonify({'error': str(exc)}), 403

    for campo, valor in payload.items():
        setattr(reg, campo, valor)

    try:
        _desvincular_prefacturas_de_atencion(reg)
        db.session.flush()
        _sincronizar_cruce_prefacturas_con_atencion(reg)
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        logger.error('Error editando atencion dia %s: %s', reg_id, exc)
        return jsonify({'error': 'No se pudo guardar el cambio'}), 500

    return jsonify({'registro': _serialize_atencion_dia(reg)}), 200

    reg  = AtencionDiaDetalle.query.get_or_404(reg_id)
    data = request.get_json() or {}

    # Campos editables
    campos = [
        'nro_orden', 'nro_factura', 'servicio', 'nro_identificacion',
        'nombre_paciente', 'acuerdo_comercial', 'empresa_mision',
        'sede', 'nombre_vendedor', 'estado_orden', 'forma_pago',
    ]
    for campo in campos:
        if campo in data:
            setattr(reg, campo, (_normalizar(data[campo]) if data[campo] not in (None, '') else None))

    if 'precio' in data:
        reg.precio = _parse_precio(data['precio'])

    if 'fecha_creacion_orden' in data and data['fecha_creacion_orden']:
        try:
            reg.fecha_creacion_orden = datetime.strptime(
                str(data['fecha_creacion_orden']).strip(), '%Y-%m-%d'
            )
        except ValueError:
            return jsonify({'error': 'Formato de fecha_creacion_orden inválido (YYYY-MM-DD)'}), 400

    if 'fecha_factura' in data and data['fecha_factura']:
        try:
            reg.fecha_factura = datetime.strptime(
                str(data['fecha_factura']).strip(), '%Y-%m-%d'
            )
        except ValueError:
            return jsonify({'error': 'Formato de fecha_factura inválido (YYYY-MM-DD)'}), 400

    try:
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        logger.error('Error editando atencion dia %s: %s', reg_id, exc)
        return jsonify({'error': 'No se pudo guardar el cambio'}), 500

    return jsonify({'registro': _serialize_atencion_dia(reg)}), 200


# ---------------------------------------------------------------------------
# POST /api/comercial/prefacturas/regenerar-empresa
# Regenera el Excel de una empresa específica para un periodo y lo descarga
# ---------------------------------------------------------------------------
@comercial_bp.route('/atenciones-dia/<int:reg_id>', methods=['DELETE'])
@login_required
def eliminar_atencion_dia(reg_id):
    try:
        _require_commercial_permission(PERMISO_ELIMINACION_ATENCIONES)
    except PermissionError as exc:
        return jsonify({'error': str(exc)}), 403

    vendedor_scope = _resolver_vendedor_usuario_actual()
    if not _is_admin_user() and vendedor_scope is None:
        return jsonify({'error': 'No tienes un vendedor asociado'}), 403

    reg = AtencionDiaDetalle.query.get_or_404(reg_id)
    try:
        _asegurar_acceso_registro_atencion(reg, vendedor_scope)
        if (reg.estado_gestion or '').upper() != 'CARGADA':
            return jsonify({'error': 'Solo se pueden eliminar atenciones en estado CARGADA'}), 409
    except PermissionError as exc:
        return jsonify({'error': str(exc)}), 403

    cargue = reg.cargue

    try:
        _desvincular_prefacturas_de_atencion(reg)
        db.session.delete(reg)
        db.session.flush()

        if cargue is not None and cargue.nombre_archivo == NOMBRE_CARGUE_MANUAL:
            restantes = AtencionDiaDetalle.query.filter_by(cargue_id=cargue.id).count()
            if restantes == 0:
                db.session.delete(cargue)
            else:
                cargue.total_filas = restantes
                cargue.filas_importadas = restantes
                cargue.filas_duplicadas = 0
                cargue.filas_error = 0

        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        logger.error('Error eliminando atencion dia %s: %s', reg_id, exc)
        return jsonify({'error': 'No se pudo eliminar la atencion'}), 500

    return jsonify({'mensaje': 'Atencion eliminada'}), 200


@comercial_bp.route('/prefacturas/regenerar-empresa', methods=['GET'])
@login_required
def regenerar_prefactura_empresa():
    """Regenera el Excel de prefactura para una empresa y periodo específicos.
    Parámetros: empresa (nombre), fecha_desde, fecha_hasta
    Actualiza el registro BORRADOR en BD y devuelve el ZIP con los archivos de esa empresa.
    """
    try:
        _require_commercial_permission(PERMISO_CONSULTA_ATENCIONES)
    except PermissionError as exc:
        return jsonify({'error': str(exc)}), 403

    empresa_nombre = (request.args.get('empresa') or '').strip()
    fecha_desde_str = (request.args.get('fecha_desde') or '').strip()
    fecha_hasta_str = (request.args.get('fecha_hasta') or '').strip()

    if not empresa_nombre or not fecha_desde_str or not fecha_hasta_str:
        return jsonify({'error': 'Se requieren empresa, fecha_desde y fecha_hasta'}), 400

    try:
        fecha_desde = datetime.strptime(fecha_desde_str, '%Y-%m-%d')
        fecha_hasta = datetime.strptime(f'{fecha_hasta_str} 23:59:59', '%Y-%m-%d %H:%M:%S')
    except ValueError:
        return jsonify({'error': 'Formato de fecha inválido (YYYY-MM-DD)'}), 400

    vendedor_scope = _resolver_vendedor_usuario_actual()
    if not _is_admin_user() and vendedor_scope is None:
        return jsonify({'error': 'No tienes un vendedor asociado'}), 403

    # Traer registros de esa empresa en el periodo
    query = (
        AtencionDiaDetalle.query
        .outerjoin(AtencionDiaDetalle.cliente)
        .filter(
            AtencionDiaDetalle.fecha_creacion_orden >= fecha_desde,
            AtencionDiaDetalle.fecha_creacion_orden <= fecha_hasta,
        )
    )
    if not _is_admin_user():
        query = query.filter(AtencionDiaDetalle.vendedor_id == vendedor_scope.id)

    todos = query.all()
    catalogo_lookup = _construir_lookup_catalogo()

    def _nombre_empresa_reg(reg):
        if reg.cliente:
            return reg.cliente.razon_social or reg.acuerdo_comercial or 'SIN_EMPRESA'
        return reg.acuerdo_comercial or reg.empresa_mision or 'SIN_EMPRESA'

    def _bucket_forma(valor):
        norm = _normalizar_forma_pago(valor)
        if norm == 'CREDITO':
            return 'CREDITO'
        if norm in ('EFECTIVO', 'CONTADO', 'PARTICULAR', 'PARTICULARES'):
            return 'EFECTIVO'
        return None

    # Filtrar registros de la empresa solicitada
    buckets: dict = defaultdict(list)
    for reg in todos:
        if _nombre_empresa_reg(reg).upper() != empresa_nombre.upper():
            continue
        bucket = _bucket_forma(reg.forma_pago)
        if bucket is None:
            continue
        estado_norm = (reg.estado_orden or '').upper().strip()
        if estado_norm == 'ANULADA':
            continue
        tipo_servicio = _clasificar_servicio(reg.servicio, catalogo_lookup)
        if tipo_servicio == 'ECOBABY':
            continue
        if reg.servicio and 'ECOBABY' in reg.servicio.upper():
            continue
        buckets[bucket].append(reg)

    if not buckets:
        return jsonify({'error': f'No se encontraron atenciones para {empresa_nombre} en el periodo indicado'}), 404

    periodo       = f"{fecha_desde.strftime('%d%m%Y')}-{fecha_hasta.strftime('%d%m%Y')}"
    periodo_label = f"{fecha_desde.strftime('%d/%m/%Y')} al {fecha_hasta.strftime('%d/%m/%Y')}"

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    tiene_cred = bool(buckets.get('CREDITO'))
    tiene_efec = bool(buckets.get('EFECTIVO'))

    if tiene_cred and tiene_efec:
        prefijo   = 'mixto'
        secciones = [('-- CREDITO --', buckets['CREDITO']), ('-- EFECTIVO --', buckets['EFECTIVO'])]
        forma_bd  = 'MIXTO'
        regs_bd   = buckets['CREDITO'] + buckets['EFECTIVO']
    elif tiene_cred:
        prefijo   = 'cred'
        secciones = [('CREDITO', buckets['CREDITO'])]
        forma_bd  = 'CREDITO'
        regs_bd   = buckets['CREDITO']
    else:
        prefijo   = 'efec'
        secciones = [('EFECTIVO', buckets['EFECTIVO'])]
        forma_bd  = 'EFECTIVO'
        regs_bd   = buckets['EFECTIVO']

    # Reutilizar _construir_workbook definida en generar_prefacturas no es posible
    # (es una función local), así que llamamos directamente a generar_prefacturas
    # redirigiendo los parámetros. En su lugar construimos el workbook aquí.
    # Importar helpers de estilos
    font_titulo        = Font(name='Calibri', bold=True, size=14, color='FFFFFFFF')
    font_subtitulo     = Font(name='Calibri', bold=True, size=11, color='FFFFFFFF')
    font_header        = Font(name='Calibri', bold=True, size=10, color='FFFFFFFF')
    font_normal        = Font(name='Calibri', size=10)
    font_total_empresa = Font(name='Calibri', bold=True, size=11, color='FF1F4E79')
    fill_titulo   = PatternFill('solid', fgColor='FF1F4E79')
    fill_azul     = PatternFill('solid', fgColor='FF2E75B6')
    fill_total    = PatternFill('solid', fgColor='FFBDD7EE')
    thin        = Side(style='thin', color='FF9DC3E6')
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    center   = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_al  = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    right_al = Alignment(horizontal='right',  vertical='center')

    def _prep_filas(filas_regs):
        om = defaultdict(list)
        for r in filas_regs:
            k = (r.nro_identificacion or '', r.nro_orden or f'_sin_{r.id}')
            om[k].append(r)
        fd = []
        for (nid, _), regs in om.items():
            fechas = [r.fecha_creacion_orden for r in regs if r.fecha_creacion_orden]
            fs = min(fechas).strftime('%d/%m/%Y') if fechas else ''
            es = _construir_examenes_str(regs, catalogo_lookup)
            v  = sum(float(r.precio) for r in regs if r.precio is not None)
            np = (regs[0].nombre_paciente or '').strip()
            fd.append((v, es, fs, nid, np))
        fd.sort(key=lambda x: (x[0], x[1], x[2]))
        return fd

    def _calc_grupos(fd):
        grupos = []
        i = 0
        while i < len(fd):
            vg, eg = fd[i][0], fd[i][1]
            gf = []
            while i < len(fd) and fd[i][0] == vg and fd[i][1] == eg:
                gf.append(fd[i]); i += 1
            grupos.append((vg, eg, gf))
        return grupos

    def _cab(ws, titulo, sub, hdrs, n):
        cl = chr(ord('A') + n - 1)
        ws.merge_cells(f'A1:{cl}1')
        c = ws['A1']; c.value = titulo; c.font = font_titulo
        c.fill = fill_titulo; c.alignment = center; ws.row_dimensions[1].height = 28
        ws.merge_cells(f'A2:{cl}2')
        c = ws['A2']; c.value = sub; c.font = font_subtitulo
        c.fill = fill_azul; c.alignment = center; ws.row_dimensions[2].height = 20
        for ci, h in enumerate(hdrs, start=1):
            c = ws.cell(row=3, column=ci, value=h)
            c.font = font_header; c.fill = fill_azul
            c.alignment = center; c.border = border_thin
        ws.row_dimensions[3].height = 18

    wb = openpyxl.Workbook()
    ws_rel = wb.active; ws_rel.title = 'relacion-pacientes'
    ws_pf  = wb.create_sheet(title='prefactura')
    _cab(ws_rel, empresa_nombre.upper(), f'RELACION DE PACIENTES  |  Periodo: {periodo_label}',
         ['Fecha Atencion', 'ID Paciente', 'Paciente', 'Examenes', 'Valor'], 5)
    _cab(ws_pf, empresa_nombre.upper(), f'PREFACTURA  |  Periodo: {periodo_label}',
         ['Examenes', 'Cant. Pacientes', 'Valor Unit.', 'Total'], 4)

    fila_rel = 4; fila_pf = 4; total_emp = 0.0

    for label, filas_regs in secciones:
        fd     = _prep_filas(filas_regs)
        grupos = _calc_grupos(fd)
        total  = sum(vg * len(gf) for vg, _, gf in grupos)
        total_emp += total

        # Separador si no es primera seccion
        if fila_rel > 4:
            ws_rel.merge_cells(f'A{fila_rel}:E{fila_rel}')
            c = ws_rel.cell(row=fila_rel, column=1, value=label)
            c.font = font_subtitulo; c.fill = fill_azul
            c.alignment = center; c.border = border_thin
            ws_rel.row_dimensions[fila_rel].height = 20; fila_rel += 1
            ws_pf.merge_cells(f'A{fila_pf}:D{fila_pf}')
            c = ws_pf.cell(row=fila_pf, column=1, value=label)
            c.font = font_subtitulo; c.fill = fill_azul
            c.alignment = center; c.border = border_thin
            ws_pf.row_dimensions[fila_pf].height = 20; fila_pf += 1

        for (v, es, fs, nid, np) in fd:
            for ci, val in enumerate([fs, nid, np, es, v], start=1):
                c = ws_rel.cell(row=fila_rel, column=ci, value=val)
                c.font = font_normal; c.border = border_thin
                if ci in (1, 2): c.alignment = center
                elif ci == 5: c.alignment = right_al; c.number_format = '#,##0.00'
                else: c.alignment = left_al
            fila_rel += 1

        for (vg, eg, gf) in grupos:
            for ci, val in enumerate([eg, len(gf), vg, vg * len(gf)], start=1):
                c = ws_pf.cell(row=fila_pf, column=ci, value=val)
                c.font = font_normal; c.border = border_thin
                if ci == 1: c.alignment = left_al
                elif ci == 2: c.alignment = center
                else: c.alignment = right_al; c.number_format = '#,##0.00'
            fila_pf += 1

        # Total seccion
        ws_rel.merge_cells(f'A{fila_rel}:D{fila_rel}')
        c = ws_rel.cell(row=fila_rel, column=1, value='TOTAL')
        c.font = font_total_empresa; c.fill = fill_total
        c.alignment = right_al; c.border = border_thin
        c2 = ws_rel.cell(row=fila_rel, column=5, value=total)
        c2.font = font_total_empresa; c2.fill = fill_total
        c2.alignment = right_al; c2.number_format = '#,##0.00'; c2.border = border_thin
        fila_rel += 1

        ws_pf.merge_cells(f'A{fila_pf}:C{fila_pf}')
        c = ws_pf.cell(row=fila_pf, column=1, value='TOTAL')
        c.font = font_total_empresa; c.fill = fill_total
        c.alignment = right_al; c.border = border_thin
        c2 = ws_pf.cell(row=fila_pf, column=4, value=total)
        c2.font = font_total_empresa; c2.fill = fill_total
        c2.alignment = right_al; c2.number_format = '#,##0.00'; c2.border = border_thin
        fila_pf += 1

    ws_rel.column_dimensions['A'].width = 16; ws_rel.column_dimensions['B'].width = 16
    ws_rel.column_dimensions['C'].width = 32; ws_rel.column_dimensions['D'].width = 70
    ws_rel.column_dimensions['E'].width = 16
    ws_pf.column_dimensions['A'].width = 70; ws_pf.column_dimensions['B'].width = 18
    ws_pf.column_dimensions['C'].width = 16; ws_pf.column_dimensions['D'].width = 16

    # Actualizar prefactura BORRADOR en BD
    pref = PrefacturaComercial.query.filter_by(
        nombre_empresa=empresa_nombre,
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta,
        forma_pago=forma_bd,
        estado='BORRADOR',
    ).first()
    if pref:
        pacs = len({(r.nro_identificacion or '', r.nombre_paciente or '') for r in regs_bd})
        val  = sum(float(r.precio) for r in regs_bd if r.precio is not None)
        pref.cant_pacientes  = pacs
        pref.valor_total     = val
        pref.usuario_genera_id = current_user.id
        try:
            db.session.commit()
        except Exception as exc_bd:
            db.session.rollback()
            logger.warning('No se pudo actualizar prefactura en BD: %s', exc_bd)

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    nombre_safe = re.sub(r'[^\w\s\-]', '', empresa_nombre).strip()
    nombre_safe = re.sub(r'\s+', '_', nombre_safe)
    nombre_zip  = f'{prefijo}-{nombre_safe}-{periodo}.zip'

    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(f'{prefijo}-{nombre_safe}-{periodo}.xlsx', buf.read())
    zip_buf.seek(0)

    return send_file(
        zip_buf,
        mimetype='application/zip',
        as_attachment=True,
        download_name=nombre_zip,
    )

# Orden de clasificaciÃÂ³n para agrupar exÃÂ¡menes en la prefactura
_ORDEN_CLASIFICACION = ['CONSULTA', 'PARACLINICO', 'LABORATORIO', 'CURSOS', 'OTRO']

# Etiquetas legibles por clasificaciÃÂ³n
_LABEL_CLASIFICACION = {
    'CONSULTA': 'Consultas',
    'PARACLINICO': 'Paraclinicos',
    'LABORATORIO': 'Laboratorios',
    'CURSOS': 'Cursos',
    'OTRO': 'Otros',
}


def _construir_lookup_catalogo():
    """Devuelve dict nombre_upper -> dict con tipo_examen y nombre_corto."""
    items = ComercialCatalogoItem.query.all()
    lookup = {}
    for item in items:
        if item.nombre:
            lookup[item.nombre.upper().strip()] = {
                'tipo_examen': (item.tipo_examen or 'OTRO').upper(),
                'nombre_corto': (item.nombre_corto or '').strip() or None,
            }
    return lookup


def _clasificar_servicio(nombre_servicio, catalogo_lookup):
    """Retorna el tipo_examen del servicio segun el catalogo, o OTRO."""
    if not nombre_servicio:
        return 'OTRO'
    entrada = catalogo_lookup.get(nombre_servicio.upper().strip())
    if entrada is None:
        return 'OTRO'
    tipo = entrada['tipo_examen']
    if tipo in ('CONSULTA', 'PARACLINICO', 'LABORATORIO', 'CURSOS', 'ECOBABY'):
        return tipo
    return 'OTRO'


def _abreviar_servicio(nombre_servicio, tipo_examen, catalogo_lookup):
    """Aplica la regla de abreviacion segun la clasificacion:
    - Si el item tiene nombre_corto en el catalogo -> usar ese.
    - CONSULTA  -> nombre completo.
    - PARACLINICO / LABORATORIO -> primeras 5 letras (mayusculas).
    - CURSOS / OTRO -> nombre completo.
    """
    if not nombre_servicio:
        return ''
    entrada = catalogo_lookup.get(nombre_servicio.upper().strip())
    if entrada and entrada.get('nombre_corto'):
        return entrada['nombre_corto']
    if tipo_examen in ('PARACLINICO', 'LABORATORIO'):
        return nombre_servicio.strip()[:5].upper()
    return nombre_servicio.strip()

def _construir_examenes_str(regs_orden, catalogo_lookup):
    """Agrupa los servicios de una orden por clasificaciÃÂ³n y construye
    la cadena de exÃÂ¡menes con el formato:
      Consultas: <nombre completo> - <nombre completo> | ParaclÃÂ­nicos: <5let> - <5let> | ...
    """
    from collections import defaultdict
    grupos = defaultdict(list)
    for reg in regs_orden:
        if not reg.servicio:
            continue
        tipo = _clasificar_servicio(reg.servicio, catalogo_lookup)
        abrev = _abreviar_servicio(reg.servicio, tipo, catalogo_lookup)
        grupos[tipo].append(abrev)

    partes = []
    for clasificacion in _ORDEN_CLASIFICACION:
        items_grupo = grupos.get(clasificacion)
        if not items_grupo:
            continue
        label = _LABEL_CLASIFICACION.get(clasificacion, clasificacion.capitalize())
        # Ordenar alfabéticamente para que el mismo conjunto de exámenes
        # siempre produzca la misma cadena sin importar el orden de llegada
        items_ordenados = sorted(items_grupo)
        partes.append(f"{label}: {' - '.join(items_ordenados)}")

    return ' | '.join(partes) if partes else ''


def _normalizar_forma_pago(valor):
    """Normaliza forma_pago eliminando tildes para comparar con 'CREDITO'."""
    if not valor:
        return ''
    import unicodedata as _ud
    texto = _ud.normalize('NFKD', str(valor))
    texto = ''.join(ch for ch in texto if not _ud.combining(ch))
    return texto.upper().strip()


@comercial_bp.route('/prefacturas/generar', methods=['GET'])
@login_required
def generar_prefacturas():
    """Genera prefacturas en Excel por empresa para un rango de fechas.

    Reglas de negocio:
    - Solo ordenes con forma_pago == 'CREDITO' (normalizado sin tilde).
    - Excluir servicios cuyo tipo_examen en catalogo sea 'ECOBABY'.
    - Agrupar por empresa -> paciente -> nro_orden.
    - Dentro de cada orden, agrupar examenes por clasificacion y abreviar
      segun tipo: CONSULTA=nombre completo, PARACLINICO/LABORATORIO=5 letras.
    - Un archivo Excel por empresa, nombre: <RazonSocial>-<periodo>.xlsx
    - Todos los archivos se guardan en la carpeta indicada por el parametro
      'carpeta' y ademas se devuelven en un ZIP de descarga.
    """
    try:
        _require_commercial_permission(PERMISO_CONSULTA_ATENCIONES)
    except PermissionError as exc:
        return jsonify({'error': str(exc)}), 403

    fecha_desde_str = request.args.get('fecha_desde', '').strip()
    fecha_hasta_str = request.args.get('fecha_hasta', '').strip()

    if not fecha_desde_str or not fecha_hasta_str:
        return jsonify({'error': 'Se requieren fecha_desde y fecha_hasta (YYYY-MM-DD)'}), 400

    try:
        fecha_desde = datetime.strptime(fecha_desde_str, '%Y-%m-%d')
        fecha_hasta = datetime.strptime(f'{fecha_hasta_str} 23:59:59', '%Y-%m-%d %H:%M:%S')
    except ValueError:
        return jsonify({'error': 'Formato de fecha invalido. Use YYYY-MM-DD'}), 400

    vendedor_scope = _resolver_vendedor_usuario_actual()
    if not _is_admin_user() and vendedor_scope is None:
        return jsonify({'error': 'No tienes un vendedor asociado para generar prefacturas'}), 403

    # -----------------------------------------------------------------------
    # Consulta base: CREDITO / CONTADO / EFECTIVO / PARTICULAR, sin ECOBABY
    # -----------------------------------------------------------------------
    query = (
        AtencionDiaDetalle.query
        .outerjoin(AtencionDiaDetalle.cliente)
        .filter(
            AtencionDiaDetalle.fecha_creacion_orden >= fecha_desde,
            AtencionDiaDetalle.fecha_creacion_orden <= fecha_hasta,
        )
    )

    if not _is_admin_user():
        query = query.filter(AtencionDiaDetalle.vendedor_id == vendedor_scope.id)

    todos = query.order_by(
        AtencionDiaDetalle.cliente_id.asc().nullslast(),
        AtencionDiaDetalle.acuerdo_comercial.asc().nullslast(),
        AtencionDiaDetalle.fecha_creacion_orden.asc().nullslast(),
        AtencionDiaDetalle.nro_identificacion.asc().nullslast(),
        AtencionDiaDetalle.nro_orden.asc().nullslast(),
    ).all()

    catalogo_lookup = _construir_lookup_catalogo()

    def _bucket_forma(valor):
        """Clasifica la forma de pago en 'CREDITO' o 'EFECTIVO' (bucket)."""
        norm = _normalizar_forma_pago(valor)
        if norm == 'CREDITO':
            return 'CREDITO'
        if norm in ('EFECTIVO', 'CONTADO', 'PARTICULAR', 'PARTICULARES'):
            return 'EFECTIVO'
        return None   # ignorar el resto

    def _nombre_empresa(reg):
        if reg.cliente:
            return reg.cliente.razon_social or reg.acuerdo_comercial or 'SIN_EMPRESA'
        return reg.acuerdo_comercial or reg.empresa_mision or 'SIN_EMPRESA'

    # Agrupar: empresa -> bucket -> [registros]
    # empresa_buckets: { nombre_empresa -> { 'CREDITO': [...], 'EFECTIVO': [...] } }
    empresa_buckets: dict = defaultdict(lambda: defaultdict(list))
    for reg in todos:
        bucket = _bucket_forma(reg.forma_pago)
        if bucket is None:
            continue
        estado_norm = (reg.estado_orden or '').upper().strip()
        if estado_norm == 'ANULADA':
            continue
        tipo_servicio = _clasificar_servicio(reg.servicio, catalogo_lookup)
        if tipo_servicio == 'ECOBABY':
            continue
        if reg.servicio and 'ECOBABY' in reg.servicio.upper():
            continue
        empresa_buckets[_nombre_empresa(reg)][bucket].append(reg)

    if not empresa_buckets:
        return jsonify({'error': 'No se encontraron atenciones en el rango de fechas indicado'}), 404

    # Para el resumen seguimos usando solo CREDITO
    empresas = {
        emp: buckets['CREDITO']
        for emp, buckets in empresa_buckets.items()
        if buckets.get('CREDITO')
    }

    periodo      = f"{fecha_desde.strftime('%d%m%Y')}-{fecha_hasta.strftime('%d%m%Y')}"
    periodo_label = f"{fecha_desde.strftime('%d/%m/%Y')} al {fecha_hasta.strftime('%d/%m/%Y')}"

    # -----------------------------------------------------------------------
    # Construir Excel por empresa
    # Reglas de prefijo y contenido:
    #   - Solo CREDITO          -> cred-Empresa.xlsx   (1 seccion: credito)
    #   - Solo EFECTIVO         -> efec-Empresa.xlsx   (1 seccion: efectivo)
    #   - CREDITO + EFECTIVO    -> mixto-Empresa.xlsx  (2 secciones: credito arriba, efectivo abajo)
    # Cada archivo tiene 2 hojas: relacion-pacientes y prefactura
    # -----------------------------------------------------------------------
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    archivos: list = []

    # ---- Estilos (definidos una vez, reutilizados en todas las funciones) ----
    def _make_styles():
        return dict(
            font_titulo        = Font(name='Calibri', bold=True, size=14, color='FFFFFFFF'),
            font_subtitulo     = Font(name='Calibri', bold=True, size=11, color='FFFFFFFF'),
            font_header        = Font(name='Calibri', bold=True, size=10, color='FFFFFFFF'),
            font_normal        = Font(name='Calibri', size=10),
            font_total_empresa = Font(name='Calibri', bold=True, size=11, color='FF1F4E79'),
            fill_titulo        = PatternFill('solid', fgColor='FF1F4E79'),
            fill_azul          = PatternFill('solid', fgColor='FF2E75B6'),
            fill_total         = PatternFill('solid', fgColor='FFBDD7EE'),
            thin               = Side(style='thin', color='FF9DC3E6'),
            center             = Alignment(horizontal='center', vertical='center', wrap_text=True),
            left_al            = Alignment(horizontal='left',   vertical='center', wrap_text=True),
            right_al           = Alignment(horizontal='right',  vertical='center'),
        )

    def _preparar_filas(filas_regs):
        """Convierte registros en lista de (valor, examenes_str, fecha_str, nro_id, nombre_pac)
        ordenada por valor asc -> examenes asc -> fecha asc."""
        ordenes_map = defaultdict(list)
        for reg in filas_regs:
            key_ord = (reg.nro_identificacion or '', reg.nro_orden or f'_sin_{reg.id}')
            ordenes_map[key_ord].append(reg)

        filas_detalle = []
        for (nro_id, _), regs_orden in ordenes_map.items():
            fechas       = [r.fecha_creacion_orden for r in regs_orden if r.fecha_creacion_orden]
            fecha_str    = min(fechas).strftime('%d/%m/%Y') if fechas else ''
            examenes_str = _construir_examenes_str(regs_orden, catalogo_lookup)
            valor        = sum(float(r.precio) for r in regs_orden if r.precio is not None)
            nombre_pac   = (regs_orden[0].nombre_paciente or '').strip()
            filas_detalle.append((valor, examenes_str, fecha_str, nro_id, nombre_pac))

        filas_detalle.sort(key=lambda x: (x[0], x[1], x[2]))
        return filas_detalle

    def _calcular_grupos(filas_detalle):
        """Agrupa filas por (valor, examenes) para la hoja prefactura."""
        grupos = []
        i = 0
        while i < len(filas_detalle):
            vg = filas_detalle[i][0]
            eg = filas_detalle[i][1]
            gf = []
            while i < len(filas_detalle) and filas_detalle[i][0] == vg and filas_detalle[i][1] == eg:
                gf.append(filas_detalle[i])
                i += 1
            grupos.append((vg, eg, gf))
        return grupos

    def _escribir_cabecera_ws(ws, titulo, subtitulo, headers_list, st):
        n = len(headers_list)
        col_fin = chr(ord('A') + n - 1)
        ws.merge_cells(f'A1:{col_fin}1')
        c = ws['A1']
        c.value = titulo; c.font = st['font_titulo']
        c.fill = st['fill_titulo']; c.alignment = st['center']
        ws.row_dimensions[1].height = 28

        ws.merge_cells(f'A2:{col_fin}2')
        c = ws['A2']
        c.value = subtitulo; c.font = st['font_subtitulo']
        c.fill = st['fill_azul']; c.alignment = st['center']
        ws.row_dimensions[2].height = 20

        border = Border(left=st['thin'], right=st['thin'], top=st['thin'], bottom=st['thin'])
        for ci, h in enumerate(headers_list, start=1):
            c = ws.cell(row=3, column=ci, value=h)
            c.font = st['font_header']; c.fill = st['fill_azul']
            c.alignment = st['center']; c.border = border
        ws.row_dimensions[3].height = 18
        return border

    def _escribir_seccion_relacion(ws, fila_inicio, filas_detalle, total, label_seccion, st):
        """Escribe una seccion de detalle en la hoja relacion-pacientes.
        Devuelve la siguiente fila disponible."""
        border = Border(left=st['thin'], right=st['thin'], top=st['thin'], bottom=st['thin'])

        # Separador de seccion si no es la primera (fila_inicio > 4)
        if fila_inicio > 4:
            fill_sep = PatternFill('solid', fgColor='FF1F4E79')
            ws.merge_cells(f'A{fila_inicio}:E{fila_inicio}')
            c = ws.cell(row=fila_inicio, column=1, value=label_seccion)
            c.font = st['font_subtitulo']; c.fill = fill_sep
            c.alignment = st['center']; c.border = border
            ws.row_dimensions[fila_inicio].height = 20
            fila_inicio += 1

        for (valor, examenes_str, fecha_str, nro_id, nombre_pac) in filas_detalle:
            for ci, v in enumerate([fecha_str, nro_id, nombre_pac, examenes_str, valor], start=1):
                c = ws.cell(row=fila_inicio, column=ci, value=v)
                c.font = st['font_normal']; c.border = border
                if ci in (1, 2):
                    c.alignment = st['center']
                elif ci == 5:
                    c.alignment = st['right_al']; c.number_format = '#,##0.00'
                else:
                    c.alignment = st['left_al']
            fila_inicio += 1

        # Total de seccion
        ws.merge_cells(f'A{fila_inicio}:D{fila_inicio}')
        c = ws.cell(row=fila_inicio, column=1, value='TOTAL')
        c.font = st['font_total_empresa']; c.fill = st['fill_total']
        c.alignment = st['right_al']; c.border = border
        c2 = ws.cell(row=fila_inicio, column=5, value=total)
        c2.font = st['font_total_empresa']; c2.fill = st['fill_total']
        c2.alignment = st['right_al']; c2.number_format = '#,##0.00'; c2.border = border
        return fila_inicio + 1

    def _escribir_seccion_prefactura(ws, fila_inicio, grupos, total, label_seccion, st):
        """Escribe una seccion de grupos en la hoja prefactura.
        Devuelve la siguiente fila disponible."""
        border = Border(left=st['thin'], right=st['thin'], top=st['thin'], bottom=st['thin'])

        if fila_inicio > 4:
            fill_sep = PatternFill('solid', fgColor='FF1F4E79')
            ws.merge_cells(f'A{fila_inicio}:D{fila_inicio}')
            c = ws.cell(row=fila_inicio, column=1, value=label_seccion)
            c.font = st['font_subtitulo']; c.fill = fill_sep
            c.alignment = st['center']; c.border = border
            ws.row_dimensions[fila_inicio].height = 20
            fila_inicio += 1

        for (vg, eg, gf) in grupos:
            cant = len(gf)
            for ci, v in enumerate([eg, cant, vg, vg * cant], start=1):
                c = ws.cell(row=fila_inicio, column=ci, value=v)
                c.font = st['font_normal']; c.border = border
                if ci == 1:
                    c.alignment = st['left_al']
                elif ci == 2:
                    c.alignment = st['center']
                else:
                    c.alignment = st['right_al']; c.number_format = '#,##0.00'
            fila_inicio += 1

        ws.merge_cells(f'A{fila_inicio}:C{fila_inicio}')
        c = ws.cell(row=fila_inicio, column=1, value='TOTAL')
        c.font = st['font_total_empresa']; c.fill = st['fill_total']
        c.alignment = st['right_al']; c.border = border
        c2 = ws.cell(row=fila_inicio, column=4, value=total)
        c2.font = st['font_total_empresa']; c2.fill = st['fill_total']
        c2.alignment = st['right_al']; c2.number_format = '#,##0.00'; c2.border = border
        return fila_inicio + 1

    def _construir_workbook(nombre_empresa, secciones):
        """secciones: lista de (label, filas_regs) en orden de aparicion."""
        st  = _make_styles()
        wb  = openpyxl.Workbook()
        ws_rel = wb.active
        ws_rel.title = 'relacion-pacientes'
        ws_pf  = wb.create_sheet(title='prefactura')

        _escribir_cabecera_ws(
            ws_rel, nombre_empresa.upper(),
            f'RELACION DE PACIENTES  |  Periodo: {periodo_label}',
            ['Fecha Atencion', 'ID Paciente', 'Paciente', 'Examenes', 'Valor'], st)
        _escribir_cabecera_ws(
            ws_pf, nombre_empresa.upper(),
            f'PREFACTURA  |  Periodo: {periodo_label}',
            ['Examenes', 'Cant. Pacientes', 'Valor Unit.', 'Total'], st)

        fila_rel = 4
        fila_pf  = 4

        for label, filas_regs in secciones:
            fd     = _preparar_filas(filas_regs)
            grupos = _calcular_grupos(fd)
            total  = sum(vg * len(gf) for vg, _, gf in grupos)
            fila_rel = _escribir_seccion_relacion(ws_rel, fila_rel, fd, total, label, st)
            fila_pf  = _escribir_seccion_prefactura(ws_pf, fila_pf, grupos, total, label, st)

        ws_rel.column_dimensions['A'].width = 16
        ws_rel.column_dimensions['B'].width = 16
        ws_rel.column_dimensions['C'].width = 32
        ws_rel.column_dimensions['D'].width = 70
        ws_rel.column_dimensions['E'].width = 16

        ws_pf.column_dimensions['A'].width = 70
        ws_pf.column_dimensions['B'].width = 18
        ws_pf.column_dimensions['C'].width = 16
        ws_pf.column_dimensions['D'].width = 16

        return wb

    # Generar un archivo por empresa con el prefijo correcto
    for nombre_empresa, buckets in sorted(empresa_buckets.items()):
        tiene_cred = bool(buckets.get('CREDITO'))
        tiene_efec = bool(buckets.get('EFECTIVO'))

        if tiene_cred and tiene_efec:
            prefijo   = 'mixto'
            secciones = [
                ('-- CREDITO --',  buckets['CREDITO']),
                ('-- EFECTIVO --', buckets['EFECTIVO']),
            ]
        elif tiene_cred:
            prefijo   = 'cred'
            secciones = [('CREDITO', buckets['CREDITO'])]
        else:
            prefijo   = 'efec'
            secciones = [('EFECTIVO', buckets['EFECTIVO'])]

        wb = _construir_workbook(nombre_empresa, secciones)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        nombre_safe    = re.sub(r'[^\w\s\-]', '', nombre_empresa).strip()
        nombre_safe    = re.sub(r'\s+', '_', nombre_safe)
        nombre_archivo = f'{prefijo}-{nombre_safe}-{periodo}.xlsx'
        archivos.append((nombre_archivo, buf.read()))

    # -----------------------------------------------------------------------
    # Generar resumen_periodo.xlsx
    # Secciones: CREDITO | MIXTO | EFECTIVO, cada una con sus empresas A-Z,
    # subtotal por seccion y total general al final.
    # -----------------------------------------------------------------------
    import openpyxl as _openpyxl
    from openpyxl.styles import (
        Font as _Font, PatternFill as _Fill,
        Alignment as _Align, Border as _Border, Side as _Side,
    )

    wb_res = _openpyxl.Workbook()
    ws_res = wb_res.active
    ws_res.title = 'Resumen'

    _thin    = _Side(style='thin', color='FF9DC3E6')
    _border  = _Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _center  = _Align(horizontal='center', vertical='center', wrap_text=True)
    _left    = _Align(horizontal='left',   vertical='center', wrap_text=True)
    _right   = _Align(horizontal='right',  vertical='center')

    _font_titulo   = _Font(name='Calibri', bold=True, size=13, color='FFFFFFFF')
    _font_seccion  = _Font(name='Calibri', bold=True, size=11, color='FFFFFFFF')
    _font_header   = _Font(name='Calibri', bold=True, size=10, color='FFFFFFFF')
    _font_data     = _Font(name='Calibri', size=10)
    _font_subtotal = _Font(name='Calibri', bold=True, size=10, color='FF1F4E79')
    _font_total    = _Font(name='Calibri', bold=True, size=11, color='FF1F4E79')

    _fill_titulo   = _Fill('solid', fgColor='FF1F4E79')
    _fill_azul     = _Fill('solid', fgColor='FF2E75B6')
    _fill_subtotal = _Fill('solid', fgColor='FFDAE3F3')
    _fill_total    = _Fill('solid', fgColor='FFBDD7EE')

    _encabezados = ['Empresa', 'Cantidad de Pacientes', 'Valor Total',
                    'Fecha Factura', 'Nro Factura', 'Valor Factura']

    # Fila 1: titulo general
    ws_res.merge_cells('A1:F1')
    c = ws_res['A1']
    c.value = f'RESUMEN PREFACTURAS  |  Periodo: {periodo_label}'
    c.font = _font_titulo; c.fill = _fill_titulo
    c.alignment = _center
    ws_res.row_dimensions[1].height = 26

    # Fila 2: encabezados de columna
    for _ci, _enc in enumerate(_encabezados, start=1):
        _c = ws_res.cell(row=2, column=_ci, value=_enc)
        _c.font = _font_header; _c.fill = _fill_azul
        _c.alignment = _center; _c.border = _border
    ws_res.row_dimensions[2].height = 18

    _fila_res   = 3
    _gran_pac   = 0
    _gran_val   = 0.0

    def _escribir_fila_empresa(ws, fila, emp, pacs, val):
        _row = [emp, pacs, val, '', '', '']
        for _ci, _v in enumerate(_row, start=1):
            _c = ws.cell(row=fila, column=_ci, value=_v)
            _c.font = _font_data; _c.border = _border
            if _ci == 1:
                _c.alignment = _left
            elif _ci == 2:
                _c.alignment = _center
            elif _ci == 3:
                _c.alignment = _right; _c.number_format = '#,##0.00'
            else:
                _c.alignment = _center

    def _escribir_subtotal(ws, fila, label, pacs, val):
        _c = ws.cell(row=fila, column=1, value=label)
        _c.font = _font_subtotal; _c.fill = _fill_subtotal
        _c.alignment = _right; _c.border = _border
        _c2 = ws.cell(row=fila, column=2, value=pacs)
        _c2.font = _font_subtotal; _c2.fill = _fill_subtotal
        _c2.alignment = _center; _c2.border = _border
        _c3 = ws.cell(row=fila, column=3, value=val)
        _c3.font = _font_subtotal; _c3.fill = _fill_subtotal
        _c3.alignment = _right; _c3.number_format = '#,##0.00'; _c3.border = _border
        for _ci in (4, 5, 6):
            _cx = ws.cell(row=fila, column=_ci, value='')
            _cx.fill = _fill_subtotal; _cx.border = _border

    # Construir datos por seccion: CREDITO, MIXTO, EFECTIVO
    # Una empresa es MIXTO si tiene ambos buckets, CREDITO si solo credito, EFECTIVO si solo efectivo
    _secciones_resumen = [
        ('CREDITO',  'CRÉDITO'),
        ('MIXTO',    'MIXTO'),
        ('EFECTIVO', 'EFECTIVO'),
    ]

    for _bucket_key, _label_sec in _secciones_resumen:
        # Recoger empresas que pertenecen a esta seccion
        _empresas_sec = {}
        for _emp, _buckets in sorted(empresa_buckets.items(), key=lambda x: x[0].upper()):
            _tiene_cred = bool(_buckets.get('CREDITO'))
            _tiene_efec = bool(_buckets.get('EFECTIVO'))
            if _bucket_key == 'CREDITO' and _tiene_cred and not _tiene_efec:
                _empresas_sec[_emp] = _buckets['CREDITO']
            elif _bucket_key == 'MIXTO' and _tiene_cred and _tiene_efec:
                _empresas_sec[_emp] = _buckets['CREDITO'] + _buckets['EFECTIVO']
            elif _bucket_key == 'EFECTIVO' and _tiene_efec and not _tiene_cred:
                _empresas_sec[_emp] = _buckets['EFECTIVO']

        if not _empresas_sec:
            continue

        # Fila separador de seccion
        ws_res.merge_cells(f'A{_fila_res}:F{_fila_res}')
        _cs = ws_res.cell(row=_fila_res, column=1, value=f'── {_label_sec} ──')
        _cs.font = _font_seccion; _cs.fill = _fill_azul
        _cs.alignment = _center; _cs.border = _border
        ws_res.row_dimensions[_fila_res].height = 20
        _fila_res += 1

        # Titulos de columna para esta seccion (solo MIXTO y EFECTIVO, CREDITO ya los tiene en fila 2)
        if _bucket_key != 'CREDITO':
            for _ci, _enc in enumerate(_encabezados, start=1):
                _c = ws_res.cell(row=_fila_res, column=_ci, value=_enc)
                _c.font = _font_header; _c.fill = _fill_azul
                _c.alignment = _center; _c.border = _border
            ws_res.row_dimensions[_fila_res].height = 18
            _fila_res += 1

        _sec_pac = 0
        _sec_val = 0.0

        for _emp, _regs_emp in _empresas_sec.items():
            _pacs = len({(r.nro_identificacion or '', r.nombre_paciente or '') for r in _regs_emp})
            _val  = sum(float(r.precio) for r in _regs_emp if r.precio is not None)
            _sec_pac += _pacs
            _sec_val += _val
            _escribir_fila_empresa(ws_res, _fila_res, _emp, _pacs, _val)
            _fila_res += 1

        # Subtotal de seccion
        _escribir_subtotal(ws_res, _fila_res, f'Subtotal {_label_sec}', _sec_pac, _sec_val)
        _fila_res += 1

        _gran_pac += _sec_pac
        _gran_val += _sec_val

    # Fila en blanco separadora antes del total general
    for _ci in range(1, 7):
        _c = ws_res.cell(row=_fila_res, column=_ci, value='')
        _c.border = _border
    _fila_res += 1

    # Fila total general
    _ct = ws_res.cell(row=_fila_res, column=1, value='TOTAL GENERAL')
    _ct.font = _font_total; _ct.fill = _fill_total
    _ct.alignment = _right; _ct.border = _border
    _cp = ws_res.cell(row=_fila_res, column=2, value=_gran_pac)
    _cp.font = _font_total; _cp.fill = _fill_total
    _cp.alignment = _center; _cp.border = _border
    _cv = ws_res.cell(row=_fila_res, column=3, value=_gran_val)
    _cv.font = _font_total; _cv.fill = _fill_total
    _cv.alignment = _right; _cv.number_format = '#,##0.00'; _cv.border = _border
    for _ci in (4, 5, 6):
        _c = ws_res.cell(row=_fila_res, column=_ci, value='')
        _c.fill = _fill_total; _c.border = _border

    ws_res.column_dimensions['A'].width = 45
    ws_res.column_dimensions['B'].width = 22
    ws_res.column_dimensions['C'].width = 18
    ws_res.column_dimensions['D'].width = 16
    ws_res.column_dimensions['E'].width = 18
    ws_res.column_dimensions['F'].width = 16

    _buf_res = io.BytesIO()
    wb_res.save(_buf_res)
    _buf_res.seek(0)
    archivos.append(('resumen_periodo.xlsx', _buf_res.read()))

    # -----------------------------------------------------------------------
    # Persistir / actualizar prefacturas en BD (estado BORRADOR)
    # Solo se actualizan las que están en BORRADOR; las CERRADAS no se tocan.
    # -----------------------------------------------------------------------
    try:
        for nombre_empresa, buckets in empresa_buckets.items():
            tiene_cred = bool(buckets.get('CREDITO'))
            tiene_efec = bool(buckets.get('EFECTIVO'))

            if tiene_cred and tiene_efec:
                formas_guardar = [('MIXTO', buckets['CREDITO'] + buckets['EFECTIVO'])]
            elif tiene_cred:
                formas_guardar = [('CREDITO', buckets['CREDITO'])]
            else:
                formas_guardar = [('EFECTIVO', buckets['EFECTIVO'])]

            for forma_bd, regs_bd in formas_guardar:
                pacs_bd = len({
                    (r.nro_identificacion or '', r.nombre_paciente or '')
                    for r in regs_bd
                })
                val_bd = sum(float(r.precio) for r in regs_bd if r.precio is not None)

                cliente_bd = next(
                    (r.cliente for r in regs_bd if r.cliente is not None), None
                )

                pref = PrefacturaComercial.query.filter_by(
                    nombre_empresa=nombre_empresa,
                    fecha_desde=fecha_desde,
                    fecha_hasta=fecha_hasta,
                    forma_pago=forma_bd,
                ).first()

                if pref is None:
                    pref = PrefacturaComercial(
                        nombre_empresa=nombre_empresa,
                        fecha_desde=fecha_desde,
                        fecha_hasta=fecha_hasta,
                        forma_pago=forma_bd,
                        estado='BORRADOR',
                        usuario_genera_id=current_user.id,
                    )
                    db.session.add(pref)

                # Solo actualizar si sigue en BORRADOR
                if pref.estado == 'BORRADOR':
                    pref.cliente_id = cliente_bd.id if cliente_bd else pref.cliente_id
                    pref.cant_pacientes = pacs_bd
                    pref.valor_total = val_bd
                    pref.usuario_genera_id = current_user.id

        db.session.commit()
    except (ProgrammingError, OperationalError) as exc_bd:
        db.session.rollback()
        mensaje_pref = _mensaje_error_esquema_prefacturas(exc_bd)
        if mensaje_pref:
            logger.warning(mensaje_pref)
        else:
            logger.warning('No se pudo guardar prefacturas en BD: %s', exc_bd)
    except Exception as exc_bd:
        db.session.rollback()
        logger.warning('No se pudo guardar prefacturas en BD: %s', exc_bd)

    # -----------------------------------------------------------------------
    # Respuesta: ZIP con todos los archivos
    # -----------------------------------------------------------------------
    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        for nombre_archivo, contenido in archivos:
            zf.writestr(nombre_archivo, contenido)
    zip_buf.seek(0)

    nombre_zip = f'Prefacturas-{periodo}.zip'
    return send_file(
        zip_buf,
        mimetype='application/zip',
        as_attachment=True,
        download_name=nombre_zip,
    )


# ===========================================================================
# ENDPOINTS DE PREFACTURAS (CRUD + CONSULTA + CARTERA)
# ===========================================================================

PREF_ORIGEN_ATENCIONES = 'ATENCIONES'
PREF_ORIGEN_MANUAL = 'MANUAL_ANTICIPO'
PREF_DETALLE_CRUCE_PENDIENTE = 'PENDIENTE'
PREF_DETALLE_CRUCE_CRUZADO = 'CRUZADO'
_TIPOS_MOVIMIENTO_CARTERA = {'PAGO_FACTURA', 'ANTICIPO', 'ABONO', 'NOTA_CREDITO'}
_MEDIOS_PAGO_CARTERA = {'EFECTIVO', 'TRANSFERENCIA', 'CHEQUE'}
_ESTADOS_CARTERA = {'APLICADO', 'PENDIENTE', 'ANULADO'}
_CANALES_TRANSFERENCIA_CARTERA = {'NEQUI', 'DAVIPLATA', 'BANCO'}


def _prefactura_total_pagado(prefactura: PrefacturaComercial) -> Decimal:
    total = Decimal('0')
    for pago in prefactura.pagos_cartera:
        if (pago.estado or '').upper() == 'ANULADO':
            continue
        total += Decimal(str(pago.valor_pago or 0))
    return total


def _prefactura_saldo(prefactura: PrefacturaComercial) -> Decimal:
    base = Decimal(str(prefactura.valor_factura or prefactura.valor_total or 0))
    return base - _prefactura_total_pagado(prefactura)


def _prefactura_manual_bloqueada(prefactura: PrefacturaComercial) -> bool:
    if (prefactura.origen or PREF_ORIGEN_ATENCIONES).upper() != PREF_ORIGEN_MANUAL:
        return False
    if prefactura.bloqueada_por_pago is True:
        return True
    total_programado = Decimal(str(prefactura.valor_total or 0))
    if total_programado <= 0:
        return False
    return _prefactura_total_pagado(prefactura) >= total_programado


def _actualizar_totales_prefactura_manual(prefactura: PrefacturaComercial) -> None:
    if (prefactura.origen or PREF_ORIGEN_ATENCIONES).upper() != PREF_ORIGEN_MANUAL:
        return

    detalles = prefactura.detalles.all()
    total = Decimal('0')
    pacientes = set()
    fecha_programada = None
    fecha_hasta = None
    for detalle in detalles:
        total += Decimal(str(detalle.valor_item or 0))
        pacientes.add((_normalizar(detalle.paciente_documento) or '', _normalizar(detalle.paciente_nombre) or ''))
        if detalle.fecha_programada and (fecha_programada is None or detalle.fecha_programada < fecha_programada):
            fecha_programada = detalle.fecha_programada
        if detalle.fecha_programada and (fecha_hasta is None or detalle.fecha_programada > fecha_hasta):
            fecha_hasta = detalle.fecha_programada

    prefactura.valor_total = total
    prefactura.valor_factura = total
    prefactura.cant_pacientes = len([item for item in pacientes if item != ('', '')])
    prefactura.fecha_programada = fecha_programada
    if fecha_programada is not None:
        prefactura.fecha_desde = fecha_programada
        prefactura.fecha_hasta = fecha_hasta or fecha_programada


def _actualizar_bloqueo_prefactura(prefactura: PrefacturaComercial) -> None:
    if (prefactura.origen or PREF_ORIGEN_ATENCIONES).upper() != PREF_ORIGEN_MANUAL:
        return
    if prefactura.bloqueada_por_pago is True:
        return
    total_programado = Decimal(str(prefactura.valor_total or 0))
    if total_programado > 0 and _prefactura_total_pagado(prefactura) >= total_programado:
        prefactura.bloqueada_por_pago = True
        prefactura.fecha_bloqueo_pago = prefactura.fecha_bloqueo_pago or datetime.utcnow()


def _asegurar_prefactura_manual_editable(prefactura: PrefacturaComercial) -> None:
    if (prefactura.origen or PREF_ORIGEN_ATENCIONES).upper() != PREF_ORIGEN_MANUAL:
        raise ValueError('Esta operacion solo aplica para prefacturas manuales de anticipo')
    if prefactura.estado == 'CERRADA':
        raise ValueError('La prefactura ya esta cerrada y no admite cambios')
    if _prefactura_manual_bloqueada(prefactura):
        raise ValueError('La prefactura ya quedo bloqueada porque el anticipo cubrio el total programado')


def _prefactura_detalle_to_dict(detalle: PrefacturaComercialDetalle) -> dict:
    return {
        'id': detalle.id,
        'prefactura_id': detalle.prefactura_id,
        'paciente_documento': detalle.paciente_documento,
        'paciente_nombre': detalle.paciente_nombre,
        'catalogo_item_id': detalle.catalogo_item_id,
        'tipo_item': detalle.tipo_item,
        'nombre_item': detalle.nombre_item,
        'valor_item': float(detalle.valor_item or 0),
        'fecha_programada': detalle.fecha_programada.strftime('%Y-%m-%d') if detalle.fecha_programada else None,
        'estado_cruce': detalle.estado_cruce,
        'atencion_dia_id': detalle.atencion_dia_id,
        'cruzado_at': detalle.cruzado_at.strftime('%Y-%m-%d %H:%M') if detalle.cruzado_at else None,
        'observaciones': detalle.observaciones,
    }


def _guardar_comprobante_cartera_prefactura(pago: CarteraPrefactura, archivo) -> None:
    if not archivo or not archivo.filename:
        return

    upload_root = current_app.config['UPLOAD_FOLDER']
    pago_dir = os.path.join(upload_root, 'comercial', 'prefacturas', str(pago.prefactura_id), 'cartera')
    os.makedirs(pago_dir, exist_ok=True)

    nombre_original = secure_filename(archivo.filename)
    if not nombre_original:
        return

    nombre_guardado = f'{uuid.uuid4().hex}_{nombre_original}'
    ruta_absoluta = os.path.join(pago_dir, nombre_guardado)
    archivo.save(ruta_absoluta)

    _eliminar_comprobante_cartera_prefactura(pago)

    pago.nombre_comprobante = nombre_original
    pago.ruta_comprobante = os.path.relpath(ruta_absoluta, upload_root)
    pago.mime_type = archivo.mimetype
    pago.tamano_bytes = os.path.getsize(ruta_absoluta)


def _eliminar_comprobante_cartera_prefactura(pago: CarteraPrefactura) -> None:
    if not pago.ruta_comprobante:
        return

    upload_root = os.path.abspath(current_app.config['UPLOAD_FOLDER'])
    ruta = os.path.abspath(os.path.join(upload_root, pago.ruta_comprobante))
    if ruta.startswith(upload_root) and os.path.exists(ruta):
        try:
            os.remove(ruta)
        except OSError:
            logger.warning('No se pudo eliminar el comprobante de cartera %s', ruta)

    pago.nombre_comprobante = None
    pago.ruta_comprobante = None
    pago.mime_type = None
    pago.tamano_bytes = None


def _get_comprobante_cartera_prefactura_path(pago: CarteraPrefactura) -> str:
    if not pago.ruta_comprobante:
        raise FileNotFoundError('Comprobante no encontrado')
    upload_root = os.path.abspath(current_app.config['UPLOAD_FOLDER'])
    ruta = os.path.abspath(os.path.join(upload_root, pago.ruta_comprobante))
    if not ruta.startswith(upload_root):
        raise FileNotFoundError('Ruta de comprobante invalida')
    if not os.path.exists(ruta):
        raise FileNotFoundError('Comprobante no encontrado')
    return ruta


def _cargar_convenio_cliente_para_prefactura(cliente: ClienteComercial, fecha_programada: datetime) -> dict[int, dict]:
    from app.routes.comercial import _build_cliente_convenio_items

    items = _build_cliente_convenio_items(cliente, fecha_programada)
    return {
        int(item['id']): item
        for item in items
        if item.get('id') is not None
    }


def _parsear_detalles_prefactura_manual(cliente: ClienteComercial, fecha_programada: datetime, detalles_raw) -> list[dict]:
    try:
        detalles = json.loads(detalles_raw or '[]')
    except json.JSONDecodeError as exc:
        raise ValueError('No se pudieron leer los detalles de la prefactura manual') from exc

    if not isinstance(detalles, list) or not detalles:
        raise ValueError('Debes agregar al menos un paciente con un examen o paquete convenido')

    convenio = _cargar_convenio_cliente_para_prefactura(cliente, fecha_programada)
    payload_detalles = []
    vistos = set()
    for detalle in detalles:
        if not isinstance(detalle, dict):
            raise ValueError('Cada detalle programado debe tener un formato valido')

        paciente_documento = _normalizar(detalle.get('paciente_documento'))
        paciente_nombre = _normalizar(detalle.get('paciente_nombre'))
        fecha_detalle_raw = _normalizar(detalle.get('fecha_programada'))
        try:
            item_id = int(detalle.get('catalogo_item_id') or detalle.get('item_id') or 0)
        except (TypeError, ValueError):
            item_id = 0
        try:
            fecha_detalle = datetime.strptime(fecha_detalle_raw, '%Y-%m-%d') if fecha_detalle_raw else fecha_programada
        except ValueError as exc:
            raise ValueError('Cada detalle debe tener una fecha valida con formato YYYY-MM-DD') from exc

        if not paciente_documento:
            raise ValueError('Cada detalle debe tener documento del paciente')
        if not paciente_nombre:
            raise ValueError('Cada detalle debe tener nombre del paciente')
        if item_id <= 0 or item_id not in convenio:
            raise ValueError('Uno o mas examenes o paquetes no estan disponibles en el convenio vigente del cliente')

        item = convenio[item_id]
        llave = (_normalizar_match(paciente_documento), _normalizar_match(item.get('nombre')), fecha_detalle.strftime('%Y-%m-%d'))
        if llave in vistos:
            raise ValueError('No puedes repetir el mismo examen o paquete para el mismo paciente dentro de la prefactura')
        vistos.add(llave)

        payload_detalles.append({
            'paciente_documento': paciente_documento,
            'paciente_nombre': paciente_nombre,
            'catalogo_item_id': item_id,
            'tipo_item': _normalizar(item.get('tipo_item')) or 'EXAMEN',
            'nombre_item': _normalizar(item.get('nombre')) or 'ITEM SIN NOMBRE',
            'valor_item': Decimal(str(item.get('valor_unitario') or 0)),
            'fecha_programada': fecha_detalle,
            'estado_cruce': PREF_DETALLE_CRUCE_PENDIENTE,
            'observaciones': _normalizar(detalle.get('observaciones')),
        })

    return payload_detalles


def _recalcular_prefactura_manual(prefactura: PrefacturaComercial) -> None:
    _actualizar_totales_prefactura_manual(prefactura)
    _actualizar_bloqueo_prefactura(prefactura)


def _prefactura_to_dict(p):
    """Serializa una PrefacturaComercial a dict."""
    pagos_total = float(_prefactura_total_pagado(p))
    saldo = float(_prefactura_saldo(p))
    origen = (p.origen or PREF_ORIGEN_ATENCIONES).upper()
    detalles = p.detalles.all() if origen == PREF_ORIGEN_MANUAL else []
    detalles_cruzados = len([detalle for detalle in detalles if detalle.atencion_dia_id is not None])
    return {
        'id':              p.id,
        'cliente_id':      p.cliente_id,
        'nombre_empresa':  p.nombre_empresa,
        'fecha_desde':     p.fecha_desde.strftime('%Y-%m-%d') if p.fecha_desde else None,
        'fecha_hasta':     p.fecha_hasta.strftime('%Y-%m-%d') if p.fecha_hasta else None,
        'fecha_programada': p.fecha_programada.strftime('%Y-%m-%d') if p.fecha_programada else None,
        'forma_pago':      p.forma_pago,
        'origen':          origen,
        'cant_pacientes':  p.cant_pacientes,
        'valor_total':     float(p.valor_total or 0),
        'estado':          p.estado,
        'bloqueada_por_pago': _prefactura_manual_bloqueada(p),
        'fecha_bloqueo_pago': p.fecha_bloqueo_pago.strftime('%Y-%m-%d %H:%M') if p.fecha_bloqueo_pago else None,
        'fecha_factura':   p.fecha_factura.strftime('%Y-%m-%d') if p.fecha_factura else None,
        'nro_factura':     p.nro_factura,
        'valor_factura':   float(p.valor_factura) if p.valor_factura is not None else None,
        'fecha_cierre':    p.fecha_cierre.strftime('%Y-%m-%d %H:%M') if p.fecha_cierre else None,
        'observaciones':   p.observaciones,
        'total_pagado':    pagos_total,
        'saldo_pendiente': saldo,
        'detalles_count':  len(detalles),
        'detalles_cruzados': detalles_cruzados,
        'created_at':      p.created_at.strftime('%Y-%m-%d %H:%M') if p.created_at else None,
        'updated_at':      p.updated_at.strftime('%Y-%m-%d %H:%M') if p.updated_at else None,
    }


def _pago_cartera_to_dict(pg):
    return {
        'id':              pg.id,
        'prefactura_id':   pg.prefactura_id,
        'tipo_movimiento': pg.tipo_movimiento,
        'fecha_pago':      pg.fecha_pago.strftime('%Y-%m-%d') if pg.fecha_pago else None,
        'valor_pago':      float(pg.valor_pago or 0),
        'medio_pago':      pg.medio_pago,
        'canal_transferencia': pg.canal_transferencia,
        'nro_comprobante': pg.nro_comprobante,
        'comprobante_nombre': pg.nombre_comprobante,
        'comprobante_url': f'/api/comercial/cartera/{pg.id}/comprobante' if pg.ruta_comprobante else None,
        'estado':          pg.estado,
        'observaciones':   pg.observaciones,
        'created_at':      pg.created_at.strftime('%Y-%m-%d %H:%M') if pg.created_at else None,
    }


# ---------------------------------------------------------------------------
# GET /api/comercial/prefacturas  — consulta con filtros
# ---------------------------------------------------------------------------
@comercial_bp.route('/prefacturas', methods=['GET'])
@login_required
def listar_prefacturas():
    """Consulta prefacturas con filtros opcionales:
    empresa, fecha_desde, fecha_hasta, forma_pago, estado.
    """
    try:
        _require_commercial_permission(PERMISO_CONSULTA_ATENCIONES)
    except PermissionError as exc:
        return jsonify({'error': str(exc)}), 403

    empresa    = (request.args.get('empresa') or '').strip()
    fd_str     = (request.args.get('fecha_desde') or '').strip()
    fh_str     = (request.args.get('fecha_hasta') or '').strip()
    forma      = (request.args.get('forma_pago') or '').strip().upper()
    estado     = (request.args.get('estado') or '').strip().upper()

    q = PrefacturaComercial.query

    if empresa:
        q = q.filter(PrefacturaComercial.nombre_empresa.ilike(f'%{empresa}%'))
    if fd_str:
        try:
            q = q.filter(PrefacturaComercial.fecha_desde >= datetime.strptime(fd_str, '%Y-%m-%d'))
        except ValueError:
            return jsonify({'error': 'fecha_desde invalida'}), 400
    if fh_str:
        try:
            q = q.filter(PrefacturaComercial.fecha_hasta <= datetime.strptime(f'{fh_str} 23:59:59', '%Y-%m-%d %H:%M:%S'))
        except ValueError:
            return jsonify({'error': 'fecha_hasta invalida'}), 400
    if forma:
        q = q.filter(PrefacturaComercial.forma_pago == forma)
    if estado:
        q = q.filter(PrefacturaComercial.estado == estado)

    # Scope por vendedor
    vendedor_scope = _resolver_vendedor_usuario_actual()
    if not _is_admin_user():
        if vendedor_scope is None:
            return jsonify({'prefacturas': []}), 200
        cliente_ids = [
            c.id for c in ClienteComercial.query.filter_by(vendedor_id=vendedor_scope.id).all()
        ]
        q = q.filter(PrefacturaComercial.cliente_id.in_(cliente_ids))

    prefacturas = q.order_by(
        PrefacturaComercial.fecha_desde.desc(),
        PrefacturaComercial.nombre_empresa.asc(),
    ).all()

    return jsonify({'prefacturas': [_prefactura_to_dict(p) for p in prefacturas]}), 200


# ---------------------------------------------------------------------------
# POST /api/comercial/clientes/<id>/prefacturas-manuales  — crear anticipo manual
# ---------------------------------------------------------------------------
@comercial_bp.route('/clientes/<int:cliente_id>/prefacturas-manuales', methods=['POST'])
@login_required
def crear_prefactura_manual(cliente_id):
    try:
        _require_commercial_permission(PERMISO_CARGUE_ATENCIONES)
    except PermissionError as exc:
        return jsonify({'error': str(exc)}), 403

    vendedor_scope = _resolver_vendedor_usuario_actual()
    if not _is_admin_user() and vendedor_scope is None:
        return jsonify({'error': 'No tienes un vendedor asociado'}), 403

    cliente = ClienteComercial.query.get_or_404(cliente_id)
    if not _is_admin_user() and cliente.vendedor_id != vendedor_scope.id:
        return jsonify({'error': 'No tienes acceso a este cliente'}), 403
    if cliente.condicion_comercial not in {'EFECTIVO', 'MIXTO'}:
        return jsonify({'error': 'Solo puedes crear esta prefactura manual para clientes EFECTIVO o MIXTO'}), 409

    fecha_programada_raw = _normalizar(request.form.get('fecha_programada'))
    fecha_pago_raw = _normalizar(request.form.get('fecha_pago'))
    observaciones = _normalizar(request.form.get('observaciones'))
    medio_pago = (_normalizar(request.form.get('medio_pago')) or 'EFECTIVO').upper()
    nro_comprobante = _normalizar(request.form.get('nro_comprobante'))
    canal_transferencia = (_normalizar(request.form.get('canal_transferencia')) or '').upper() or None

    if not fecha_programada_raw:
        return jsonify({'error': 'Debes indicar la fecha programada'}), 400
    if not fecha_pago_raw:
        return jsonify({'error': 'Debes indicar la fecha del anticipo'}), 400

    try:
        fecha_programada = datetime.strptime(fecha_programada_raw, '%Y-%m-%d')
        fecha_pago = datetime.strptime(fecha_pago_raw, '%Y-%m-%d')
    except ValueError:
        return jsonify({'error': 'Las fechas deben tener formato YYYY-MM-DD'}), 400

    if medio_pago not in _MEDIOS_PAGO_CARTERA:
        return jsonify({'error': 'El medio de pago debe ser EFECTIVO, TRANSFERENCIA o CHEQUE'}), 400
    if medio_pago == 'TRANSFERENCIA' and canal_transferencia not in _CANALES_TRANSFERENCIA_CARTERA:
        return jsonify({'error': 'Debes indicar si la transferencia fue por NEQUI, DAVIPLATA o BANCO'}), 400

    try:
        detalles_payload = _parsear_detalles_prefactura_manual(
            cliente,
            fecha_programada,
            request.form.get('detalles'),
        )
        valor_pago = Decimal(str(request.form.get('valor_pago') or 0))
        if valor_pago <= 0:
            raise ValueError
    except ValueError as exc:
        mensaje = str(exc) if str(exc) else 'El valor del anticipo debe ser mayor que cero'
        return jsonify({'error': mensaje}), 400

    pref_existente = PrefacturaComercial.query.filter_by(
        cliente_id=cliente.id,
        fecha_desde=fecha_programada,
        fecha_hasta=fecha_programada,
        forma_pago=cliente.condicion_comercial,
        origen=PREF_ORIGEN_MANUAL,
    ).first()
    if pref_existente is not None:
        return jsonify({'error': 'Ya existe una prefactura manual para este cliente y esta fecha programada'}), 409

    pref = PrefacturaComercial(
        cliente_id=cliente.id,
        nombre_empresa=cliente.razon_social,
        fecha_desde=fecha_programada,
        fecha_hasta=fecha_programada,
        fecha_programada=fecha_programada,
        forma_pago=cliente.condicion_comercial,
        origen=PREF_ORIGEN_MANUAL,
        estado='BORRADOR',
        observaciones=observaciones,
        usuario_genera_id=getattr(current_user, 'id', None),
    )
    db.session.add(pref)
    db.session.flush()

    for detalle_payload in detalles_payload:
        db.session.add(PrefacturaComercialDetalle(prefactura_id=pref.id, **detalle_payload))

    db.session.flush()
    _recalcular_prefactura_manual(pref)

    pago = CarteraPrefactura(
        prefactura_id=pref.id,
        tipo_movimiento='ANTICIPO',
        fecha_pago=fecha_pago,
        valor_pago=valor_pago,
        medio_pago=medio_pago,
        canal_transferencia=canal_transferencia,
        nro_comprobante=nro_comprobante,
        estado='APLICADO',
        observaciones=observaciones,
        usuario_id=getattr(current_user, 'id', None),
    )
    db.session.add(pago)
    db.session.flush()

    comprobante = request.files.get('comprobante_pago')
    if medio_pago == 'TRANSFERENCIA' and not comprobante:
        db.session.rollback()
        return jsonify({'error': 'Debes adjuntar el comprobante cuando el anticipo se registra por transferencia'}), 400
    if comprobante:
        _guardar_comprobante_cartera_prefactura(pago, comprobante)

    _actualizar_bloqueo_prefactura(pref)

    try:
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        logger.error('Error creando prefactura manual para cliente %s: %s', cliente_id, exc)
        return jsonify({'error': 'No se pudo crear la prefactura manual'}), 500

    pref = PrefacturaComercial.query.get(pref.id)
    _intentar_cruzar_prefactura_manual(pref)
    try:
        db.session.commit()
    except Exception:
        db.session.rollback()

    return jsonify({'mensaje': 'Prefactura manual creada', 'prefactura': _prefactura_to_dict(pref)}), 201


# ---------------------------------------------------------------------------
# GET /api/comercial/prefacturas/<id>  — detalle de una prefactura
# ---------------------------------------------------------------------------
@comercial_bp.route('/prefacturas/<int:pref_id>', methods=['GET'])
@login_required
def obtener_prefactura(pref_id):
    try:
        _require_commercial_permission(PERMISO_CONSULTA_ATENCIONES)
    except PermissionError as exc:
        return jsonify({'error': str(exc)}), 403

    pref = PrefacturaComercial.query.get_or_404(pref_id)
    pagos = [_pago_cartera_to_dict(pg) for pg in pref.pagos_cartera.order_by(CarteraPrefactura.fecha_pago.asc()).all()]
    data = _prefactura_to_dict(pref)
    data['pagos'] = pagos
    if (pref.origen or PREF_ORIGEN_ATENCIONES).upper() == PREF_ORIGEN_MANUAL:
        data['detalles'] = [
            _prefactura_detalle_to_dict(detalle)
            for detalle in pref.detalles.order_by(
                PrefacturaComercialDetalle.fecha_programada.asc(),
                PrefacturaComercialDetalle.paciente_nombre.asc(),
                PrefacturaComercialDetalle.nombre_item.asc(),
                PrefacturaComercialDetalle.id.asc(),
            ).all()
        ]
    return jsonify({'prefactura': data}), 200


@comercial_bp.route('/prefacturas/<int:pref_id>/detalles', methods=['POST'])
@login_required
def agregar_detalle_prefactura_manual(pref_id):
    try:
        _require_commercial_permission(PERMISO_EDICION_ATENCIONES)
    except PermissionError as exc:
        return jsonify({'error': str(exc)}), 403

    pref = PrefacturaComercial.query.get_or_404(pref_id)
    try:
        _asegurar_prefactura_manual_editable(pref)
        fecha_programada = pref.fecha_programada or pref.fecha_desde
        cliente = pref.cliente or ClienteComercial.query.get(pref.cliente_id)
        if cliente is None:
            raise ValueError('La prefactura no tiene un cliente valido asociado')
        detalles = _parsear_detalles_prefactura_manual(
            cliente,
            fecha_programada,
            json.dumps([request.get_json() or {}]),
        )
    except ValueError as exc:
        return jsonify({'error': str(exc)}), 400

    detalle = PrefacturaComercialDetalle(prefactura_id=pref.id, **detalles[0])
    db.session.add(detalle)
    db.session.flush()
    _recalcular_prefactura_manual(pref)

    try:
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        logger.error('Error agregando detalle a prefactura %s: %s', pref_id, exc)
        return jsonify({'error': 'No se pudo agregar el detalle a la prefactura'}), 500

    _intentar_cruzar_prefactura_manual(pref)
    try:
        db.session.commit()
    except Exception:
        db.session.rollback()

    return jsonify({'detalle': _prefactura_detalle_to_dict(detalle), 'prefactura': _prefactura_to_dict(pref)}), 201


@comercial_bp.route('/prefacturas/detalles/<int:detalle_id>', methods=['PUT'])
@login_required
def editar_detalle_prefactura_manual(detalle_id):
    try:
        _require_commercial_permission(PERMISO_EDICION_ATENCIONES)
    except PermissionError as exc:
        return jsonify({'error': str(exc)}), 403

    detalle = PrefacturaComercialDetalle.query.get_or_404(detalle_id)
    pref = detalle.prefactura
    try:
        _asegurar_prefactura_manual_editable(pref)
        cliente = pref.cliente or ClienteComercial.query.get(pref.cliente_id)
        if cliente is None:
            raise ValueError('La prefactura no tiene un cliente valido asociado')
        payloads = _parsear_detalles_prefactura_manual(
            cliente,
            pref.fecha_programada or pref.fecha_desde,
            json.dumps([request.get_json() or {}]),
        )
    except ValueError as exc:
        return jsonify({'error': str(exc)}), 400

    payload = payloads[0]
    detalle.paciente_documento = payload['paciente_documento']
    detalle.paciente_nombre = payload['paciente_nombre']
    detalle.catalogo_item_id = payload['catalogo_item_id']
    detalle.tipo_item = payload['tipo_item']
    detalle.nombre_item = payload['nombre_item']
    detalle.valor_item = payload['valor_item']
    detalle.fecha_programada = payload['fecha_programada']
    detalle.observaciones = payload['observaciones']
    detalle.atencion_dia_id = None
    detalle.estado_cruce = PREF_DETALLE_CRUCE_PENDIENTE
    detalle.cruzado_at = None
    _recalcular_prefactura_manual(pref)

    try:
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        logger.error('Error editando detalle manual %s: %s', detalle_id, exc)
        return jsonify({'error': 'No se pudo actualizar el detalle de la prefactura'}), 500

    _intentar_cruzar_prefactura_manual(pref)
    try:
        db.session.commit()
    except Exception:
        db.session.rollback()

    return jsonify({'detalle': _prefactura_detalle_to_dict(detalle), 'prefactura': _prefactura_to_dict(pref)}), 200


@comercial_bp.route('/prefacturas/detalles/<int:detalle_id>', methods=['DELETE'])
@login_required
def eliminar_detalle_prefactura_manual(detalle_id):
    try:
        _require_commercial_permission(PERMISO_ELIMINACION_ATENCIONES)
    except PermissionError as exc:
        return jsonify({'error': str(exc)}), 403

    detalle = PrefacturaComercialDetalle.query.get_or_404(detalle_id)
    pref = detalle.prefactura
    try:
        _asegurar_prefactura_manual_editable(pref)
    except ValueError as exc:
        return jsonify({'error': str(exc)}), 409

    try:
        db.session.delete(detalle)
        db.session.flush()
        if pref.detalles.count() == 0:
            db.session.delete(pref)
        else:
            _recalcular_prefactura_manual(pref)
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        logger.error('Error eliminando detalle manual %s: %s', detalle_id, exc)
        return jsonify({'error': 'No se pudo eliminar el detalle de la prefactura'}), 500

    return jsonify({'ok': True}), 200


# ---------------------------------------------------------------------------
# PUT /api/comercial/prefacturas/<id>  — editar (solo BORRADOR)
# Permite actualizar observaciones y datos de factura mientras esté en BORRADOR
# ---------------------------------------------------------------------------
@comercial_bp.route('/prefacturas/<int:pref_id>', methods=['PUT'])
@login_required
def actualizar_prefactura(pref_id):
    try:
        _require_commercial_permission(PERMISO_CARGUE_ATENCIONES)
    except PermissionError as exc:
        return jsonify({'error': str(exc)}), 403

    pref = PrefacturaComercial.query.get_or_404(pref_id)
    if pref.estado == 'CERRADA':
        return jsonify({'error': 'La prefactura está cerrada y no puede modificarse'}), 409

    data = request.get_json() or {}
    if (pref.origen or PREF_ORIGEN_ATENCIONES).upper() == PREF_ORIGEN_MANUAL and _prefactura_manual_bloqueada(pref):
        return jsonify({'error': 'La prefactura manual ya esta bloqueada porque el anticipo cubrio el total'}), 409
    if 'observaciones' in data:
        pref.observaciones = (data['observaciones'] or '').strip() or None

    try:
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        logger.error('Error actualizando prefactura %s: %s', pref_id, exc)
        return jsonify({'error': 'No se pudo actualizar la prefactura'}), 500

    return jsonify({'prefactura': _prefactura_to_dict(pref)}), 200


# ---------------------------------------------------------------------------
# POST /api/comercial/prefacturas/<id>/cerrar  — cerrar periodo de una prefactura
# Registra fecha_factura, nro_factura, valor_factura y pasa a CERRADA
# ---------------------------------------------------------------------------
@comercial_bp.route('/prefacturas/<int:pref_id>/cerrar', methods=['POST'])
@login_required
def cerrar_prefactura(pref_id):
    try:
        _require_commercial_permission(PERMISO_CARGUE_ATENCIONES)
    except PermissionError as exc:
        return jsonify({'error': str(exc)}), 403

    pref = PrefacturaComercial.query.get_or_404(pref_id)
    if pref.estado == 'CERRADA':
        return jsonify({'error': 'La prefactura ya está cerrada'}), 409

    data = request.get_json() or {}

    # Para crédito: fecha y nro factura son obligatorios
    if (pref.origen or PREF_ORIGEN_ATENCIONES).upper() == PREF_ORIGEN_MANUAL:
        return jsonify({'error': 'Las prefacturas manuales de anticipo no se cierran por este flujo'}), 409

    if pref.forma_pago in ('CREDITO', 'MIXTO'):
        fecha_fac_str = (data.get('fecha_factura') or '').strip()
        nro_fac       = (data.get('nro_factura') or '').strip()
        if not fecha_fac_str:
            return jsonify({'error': 'La fecha de factura es obligatoria para cerrar'}), 400
        if not nro_fac:
            return jsonify({'error': 'El número de factura es obligatorio para cerrar'}), 400
        try:
            pref.fecha_factura = datetime.strptime(fecha_fac_str, '%Y-%m-%d')
        except ValueError:
            return jsonify({'error': 'Formato de fecha_factura inválido (YYYY-MM-DD)'}), 400
        pref.nro_factura = nro_fac

    valor_fac_raw = data.get('valor_factura')
    if valor_fac_raw not in (None, ''):
        try:
            from decimal import Decimal as _Dec
            pref.valor_factura = _Dec(str(valor_fac_raw))
        except Exception:
            return jsonify({'error': 'valor_factura inválido'}), 400
    else:
        pref.valor_factura = pref.valor_total

    pref.estado           = 'CERRADA'
    pref.fecha_cierre     = datetime.utcnow()
    pref.usuario_cierra_id = current_user.id
    if data.get('observaciones'):
        pref.observaciones = str(data['observaciones']).strip()

    try:
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        logger.error('Error cerrando prefactura %s: %s', pref_id, exc)
        return jsonify({'error': 'No se pudo cerrar la prefactura'}), 500

    return jsonify({'prefactura': _prefactura_to_dict(pref)}), 200


# ---------------------------------------------------------------------------
# POST /api/comercial/prefacturas/<id>/reabrir  — reabrir a BORRADOR (admin)
# ---------------------------------------------------------------------------
@comercial_bp.route('/prefacturas/<int:pref_id>/reabrir', methods=['POST'])
@login_required
def reabrir_prefactura(pref_id):
    if not _is_admin_user():
        return jsonify({'error': 'Solo el administrador puede reabrir una prefactura cerrada'}), 403

    pref = PrefacturaComercial.query.get_or_404(pref_id)
    if pref.estado != 'CERRADA':
        return jsonify({'error': 'La prefactura no está cerrada'}), 409

    pref.estado            = 'BORRADOR'
    pref.fecha_cierre      = None
    pref.usuario_cierra_id = None

    try:
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        return jsonify({'error': 'No se pudo reabrir la prefactura'}), 500

    return jsonify({'prefactura': _prefactura_to_dict(pref)}), 200


# ---------------------------------------------------------------------------
# DELETE /api/comercial/prefacturas/<id>  — eliminar (solo BORRADOR, admin)
# ---------------------------------------------------------------------------
@comercial_bp.route('/prefacturas/<int:pref_id>', methods=['DELETE'])
@login_required
def eliminar_prefactura(pref_id):
    if not _is_admin_user():
        return jsonify({'error': 'Solo el administrador puede eliminar prefacturas'}), 403

    pref = PrefacturaComercial.query.get_or_404(pref_id)
    if pref.estado == 'CERRADA':
        return jsonify({'error': 'No se puede eliminar una prefactura cerrada'}), 409
    if (pref.origen or PREF_ORIGEN_ATENCIONES).upper() == PREF_ORIGEN_MANUAL and _prefactura_manual_bloqueada(pref):
        return jsonify({'error': 'La prefactura manual esta bloqueada y ya no se puede eliminar'}), 409

    try:
        for pago in pref.pagos_cartera.all():
            _eliminar_comprobante_cartera_prefactura(pago)
        db.session.delete(pref)
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        return jsonify({'error': 'No se pudo eliminar la prefactura'}), 500

    return jsonify({'ok': True}), 200


# ---------------------------------------------------------------------------
# POST /api/comercial/prefacturas/cargar-resumen
# Carga masiva desde resumen_periodo.xlsx: actualiza fecha/nro/valor factura
# en las prefacturas BORRADOR de crédito que coincidan por nombre de empresa
# ---------------------------------------------------------------------------
@comercial_bp.route('/prefacturas/cargar-resumen', methods=['POST'])
@login_required
def cargar_resumen_prefacturas():
    try:
        _require_commercial_permission(PERMISO_CARGUE_ATENCIONES)
    except PermissionError as exc:
        return jsonify({'error': str(exc)}), 403

    if 'archivo' not in request.files:
        return jsonify({'error': 'No se envió ningún archivo'}), 400

    archivo = request.files['archivo']
    if not (archivo.filename or '').lower().endswith('.xlsx'):
        return jsonify({'error': 'Solo se aceptan archivos .xlsx'}), 400

    import openpyxl as _xl
    try:
        wb = _xl.load_workbook(filename=io.BytesIO(archivo.read()), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    except Exception as exc:
        return jsonify({'error': f'No se pudo leer el archivo: {exc}'}), 400

    # Buscar fila de encabezados (contiene 'Empresa')
    header_row = None
    for i, row in enumerate(rows):
        vals = [str(v or '').strip().lower() for v in row]
        if 'empresa' in vals:
            header_row = i
            break

    if header_row is None:
        return jsonify({'error': 'No se encontró la fila de encabezados en el archivo'}), 400

    headers = [str(v or '').strip().lower() for v in rows[header_row]]

    def _col(name):
        for i, h in enumerate(headers):
            if name in h:
                return i
        return None

    col_empresa      = _col('empresa')
    col_fecha_fac    = _col('fecha factura')
    col_nro_fac      = _col('nro factura')
    col_valor_fac    = _col('valor factura')

    if col_empresa is None:
        return jsonify({'error': 'No se encontró la columna Empresa en el archivo'}), 400

    actualizadas = 0
    no_encontradas = []

    for row in rows[header_row + 1:]:
        if not row or all(v in (None, '') for v in row):
            continue
        nombre_emp = str(row[col_empresa] or '').strip()
        if not nombre_emp or nombre_emp.upper() in ('TOTAL', 'TOTAL GENERAL', ''):
            continue

        fecha_fac_val  = row[col_fecha_fac]  if col_fecha_fac  is not None else None
        nro_fac_val    = row[col_nro_fac]    if col_nro_fac    is not None else None
        valor_fac_val  = row[col_valor_fac]  if col_valor_fac  is not None else None

        # Parsear fecha
        fecha_fac = None
        if fecha_fac_val:
            if isinstance(fecha_fac_val, datetime):
                fecha_fac = fecha_fac_val
            else:
                for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y'):
                    try:
                        fecha_fac = datetime.strptime(str(fecha_fac_val).strip(), fmt)
                        break
                    except ValueError:
                        pass

        nro_fac = str(nro_fac_val or '').strip() or None

        valor_fac = None
        if valor_fac_val not in (None, ''):
            try:
                from decimal import Decimal as _Dec
                valor_fac = _Dec(str(valor_fac_val).replace(',', '.'))
            except Exception:
                pass

        # Buscar prefactura BORRADOR de crédito o mixto para esta empresa
        prefs = PrefacturaComercial.query.filter(
            PrefacturaComercial.nombre_empresa.ilike(nombre_emp),
            PrefacturaComercial.estado == 'BORRADOR',
            PrefacturaComercial.forma_pago.in_(['CREDITO', 'MIXTO']),
            PrefacturaComercial.origen == PREF_ORIGEN_ATENCIONES,
        ).all()

        if not prefs:
            no_encontradas.append(nombre_emp)
            continue

        for pref in prefs:
            if fecha_fac:
                pref.fecha_factura = fecha_fac
            if nro_fac:
                pref.nro_factura = nro_fac
            if valor_fac is not None:
                pref.valor_factura = valor_fac
            actualizadas += 1

    try:
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        return jsonify({'error': f'Error guardando cambios: {exc}'}), 500

    return jsonify({
        'actualizadas':    actualizadas,
        'no_encontradas':  no_encontradas,
        'mensaje': f'{actualizadas} prefactura(s) actualizadas.',
    }), 200


# ===========================================================================
# ENDPOINTS DE CARTERA (pagos sobre prefacturas)
# ===========================================================================

# ---------------------------------------------------------------------------
# GET /api/comercial/prefacturas/<id>/cartera  — pagos de una prefactura
# ---------------------------------------------------------------------------
@comercial_bp.route('/prefacturas/<int:pref_id>/cartera', methods=['GET'])
@login_required
def listar_cartera_prefactura(pref_id):
    try:
        _require_commercial_permission(PERMISO_CONSULTA_ATENCIONES)
    except PermissionError as exc:
        return jsonify({'error': str(exc)}), 403

    pref = PrefacturaComercial.query.get_or_404(pref_id)
    pagos = pref.pagos_cartera.order_by(CarteraPrefactura.fecha_pago.asc()).all()
    return jsonify({'pagos': [_pago_cartera_to_dict(pg) for pg in pagos]}), 200


@comercial_bp.route('/prefacturas/<int:pref_id>/cartera', methods=['POST'])
@login_required
def registrar_pago_cartera(pref_id):
    try:
        _require_commercial_permission(PERMISO_CARGUE_ATENCIONES)
    except PermissionError as exc:
        return jsonify({'error': str(exc)}), 403

    pref = PrefacturaComercial.query.get_or_404(pref_id)
    if (pref.origen or PREF_ORIGEN_ATENCIONES).upper() == PREF_ORIGEN_MANUAL and _prefactura_manual_bloqueada(pref):
        return jsonify({'error': 'La prefactura manual ya esta bloqueada y no admite mas movimientos'}), 409

    data = request.get_json() if request.is_json else request.form.to_dict()
    data = data or {}
    comprobante = request.files.get('comprobante_pago')

    tipo = str(data.get('tipo_movimiento') or 'PAGO_FACTURA').strip().upper()
    if tipo not in _TIPOS_MOVIMIENTO_CARTERA:
        return jsonify({'error': f'tipo_movimiento debe ser: {", ".join(_TIPOS_MOVIMIENTO_CARTERA)}'}), 400

    fecha_str = (data.get('fecha_pago') or '').strip()
    if not fecha_str:
        return jsonify({'error': 'La fecha de pago es obligatoria'}), 400
    try:
        fecha_pago = datetime.strptime(fecha_str, '%Y-%m-%d')
    except ValueError:
        return jsonify({'error': 'Formato de fecha_pago inválido (YYYY-MM-DD)'}), 400

    try:
        from decimal import Decimal as _Dec
        valor_pago = _Dec(str(data.get('valor_pago') or 0))
        if valor_pago <= 0:
            raise ValueError
    except (ValueError, Exception):
        return jsonify({'error': 'valor_pago debe ser un número mayor que cero'}), 400

    medio = str(data.get('medio_pago') or 'EFECTIVO').strip().upper()
    if medio not in _MEDIOS_PAGO_CARTERA:
        medio = 'EFECTIVO'
    canal_transferencia = (_normalizar(data.get('canal_transferencia')) or '').upper() or None
    if medio == 'TRANSFERENCIA' and canal_transferencia not in _CANALES_TRANSFERENCIA_CARTERA:
        return jsonify({'error': 'Debes indicar si la transferencia fue por NEQUI, DAVIPLATA o BANCO'}), 400

    pg = CarteraPrefactura(
        prefactura_id   = pref.id,
        tipo_movimiento = tipo,
        fecha_pago      = fecha_pago,
        valor_pago      = valor_pago,
        medio_pago      = medio,
        canal_transferencia = canal_transferencia,
        nro_comprobante = (data.get('nro_comprobante') or '').strip() or None,
        estado          = 'APLICADO',
        observaciones   = (data.get('observaciones') or '').strip() or None,
        usuario_id      = current_user.id,
    )
    db.session.add(pg)
    db.session.flush()

    if medio == 'TRANSFERENCIA' and not comprobante:
        db.session.rollback()
        return jsonify({'error': 'Debes adjuntar el comprobante cuando el pago se registra por transferencia'}), 400
    if comprobante:
        _guardar_comprobante_cartera_prefactura(pg, comprobante)
    _actualizar_bloqueo_prefactura(pref)

    try:
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        return jsonify({'error': 'No se pudo registrar el pago'}), 500

    return jsonify({'pago': _pago_cartera_to_dict(pg)}), 201


# ---------------------------------------------------------------------------
# PUT /api/comercial/cartera/<id>  — editar pago de cartera
# ---------------------------------------------------------------------------
@comercial_bp.route('/cartera/<int:pago_id>', methods=['PUT'])
@login_required
def actualizar_pago_cartera(pago_id):
    try:
        _require_commercial_permission(PERMISO_CARGUE_ATENCIONES)
    except PermissionError as exc:
        return jsonify({'error': str(exc)}), 403

    pg = CarteraPrefactura.query.get_or_404(pago_id)
    pref = pg.prefactura
    if (pref.origen or PREF_ORIGEN_ATENCIONES).upper() == PREF_ORIGEN_MANUAL and _prefactura_manual_bloqueada(pref):
        return jsonify({'error': 'La prefactura manual ya esta bloqueada y no admite cambios en sus pagos'}), 409

    data = request.get_json() if request.is_json else request.form.to_dict()
    data = data or {}
    comprobante = request.files.get('comprobante_pago')

    if 'fecha_pago' in data and data['fecha_pago']:
        try:
            pg.fecha_pago = datetime.strptime(str(data['fecha_pago']).strip(), '%Y-%m-%d')
        except ValueError:
            return jsonify({'error': 'Formato de fecha_pago inválido'}), 400

    if 'valor_pago' in data:
        try:
            from decimal import Decimal as _Dec
            v = _Dec(str(data['valor_pago']))
            if v <= 0:
                raise ValueError
            pg.valor_pago = v
        except Exception:
            return jsonify({'error': 'valor_pago inválido'}), 400

    if 'medio_pago' in data:
        medio = str(data['medio_pago'] or '').strip().upper()
        if medio in _MEDIOS_PAGO_CARTERA:
            pg.medio_pago = medio
    if 'canal_transferencia' in data:
        canal = (_normalizar(data.get('canal_transferencia')) or '').upper() or None
        if pg.medio_pago == 'TRANSFERENCIA' and canal not in _CANALES_TRANSFERENCIA_CARTERA:
            return jsonify({'error': 'Debes indicar si la transferencia fue por NEQUI, DAVIPLATA o BANCO'}), 400
        pg.canal_transferencia = canal if pg.medio_pago == 'TRANSFERENCIA' else None

    if 'nro_comprobante' in data:
        pg.nro_comprobante = (data['nro_comprobante'] or '').strip() or None

    if 'estado' in data:
        est = str(data['estado'] or '').strip().upper()
        if est in _ESTADOS_CARTERA:
            pg.estado = est

    if 'observaciones' in data:
        pg.observaciones = (data['observaciones'] or '').strip() or None
    if pg.medio_pago == 'TRANSFERENCIA' and not comprobante and not pg.ruta_comprobante:
        return jsonify({'error': 'Debes adjuntar el comprobante cuando el pago se registra por transferencia'}), 400
    if pg.medio_pago != 'TRANSFERENCIA':
        _eliminar_comprobante_cartera_prefactura(pg)
        pg.canal_transferencia = None
    elif comprobante:
        _guardar_comprobante_cartera_prefactura(pg, comprobante)

    try:
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        return jsonify({'error': 'No se pudo actualizar el pago'}), 500

    return jsonify({'pago': _pago_cartera_to_dict(pg)}), 200


# ---------------------------------------------------------------------------
# DELETE /api/comercial/cartera/<id>  — anular pago de cartera
# ---------------------------------------------------------------------------
@comercial_bp.route('/cartera/<int:pago_id>', methods=['DELETE'])
@login_required
def anular_pago_cartera(pago_id):
    try:
        _require_commercial_permission(PERMISO_CARGUE_ATENCIONES)
    except PermissionError as exc:
        return jsonify({'error': str(exc)}), 403

    pg = CarteraPrefactura.query.get_or_404(pago_id)
    pref = pg.prefactura
    if (pref.origen or PREF_ORIGEN_ATENCIONES).upper() == PREF_ORIGEN_MANUAL and _prefactura_manual_bloqueada(pref):
        return jsonify({'error': 'La prefactura manual ya esta bloqueada y no admite anulaciones de pago'}), 409
    pg.estado = 'ANULADO'

    try:
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        return jsonify({'error': 'No se pudo anular el pago'}), 500

    return jsonify({'pago': _pago_cartera_to_dict(pg)}), 200


@comercial_bp.route('/cartera/<int:pago_id>/comprobante', methods=['GET'])
@login_required
def descargar_comprobante_cartera_prefactura(pago_id):
    try:
        _require_commercial_permission(PERMISO_CONSULTA_ATENCIONES)
    except PermissionError as exc:
        return jsonify({'error': str(exc)}), 403

    pago = CarteraPrefactura.query.get_or_404(pago_id)
    try:
        ruta = _get_comprobante_cartera_prefactura_path(pago)
        return send_file(ruta, as_attachment=True, download_name=pago.nombre_comprobante or 'comprobante_pago')
    except FileNotFoundError as exc:
        return jsonify({'error': str(exc)}), 404
    except Exception as exc:
        logger.error('Error descargando comprobante de cartera %s: %s', pago_id, exc)
        return jsonify({'error': 'Error al descargar el comprobante del pago'}), 500


# ===========================================================================
# REGISTRO DIARIO DE CAJA — ÓRDENES DE SERVICIO
# ===========================================================================

_ESTADOS_ORDEN_CAJA   = {'INGRESADO', 'APROBADO', 'TERMINADO', 'ANULADO'}
_TIPOS_DOCUMENTO_CAJA = {'CC', 'CE', 'PT'}
_FORMAS_PAGO_CAJA     = {'EFECTIVO', 'TRANSFERENCIA', 'CREDITO', 'MIXTO'}
_TIPOS_CLIENTE_CAJA   = {'EMPRESA', 'PERSONA_NATURAL'}

_OPCIONES_TIPO_EXAMEN  = ['Ingreso', 'Periódico', 'Egreso', 'Post-incapacidad',
                          'Control y seguimiento', 'Reubicación']
_OPCIONES_ENFASIS      = ['Osteomuscular', 'Alturas', 'Manipulación de Alimentos',
                          'Espacios Confinados', 'Cardiovascular', 'Neurológico',
                          'Respiratorio', 'Dermatológico', 'Piel y Faneras',
                          'Manejo de alta tensión eléctrica']
_OPCIONES_PARACLINICOS = ['Audiometría', 'Visiometría', 'Optometría', 'Espirometría',
                          'Prueba Psicológica', 'Prueba Motriz', 'Electrocardiograma',
                          'Tamizaje de voz (Voximetría)']
_OPCIONES_LABORATORIO  = ['Triglicéridos', 'Glicemia Basal', 'Colesterol Total',
                          'Perfil Lipídico', 'Cuadro Hemático', 'Frótis de uñas y garganta',
                          'Coprológico', 'Hemoclasificación']
_OPCIONES_OTROS        = ['Curso de Manipulación de Alimentos', 'Curso de Alturas',
                          'Radiografía', 'Vacuna']


_OPCIONES_AUTORIZACION = ['PLATAFORMA', 'CORREO', 'ORDEN_FISICA', 'WHATSAPP_ADMISIONES']
_GRUPOS_REQUERIMIENTO = [
    {'clave': 'GERENCIA_C', 'nombre': 'Gerencia C'},
    {'clave': 'ASESORA_Y', 'nombre': 'Asesora Y'},
    {'clave': 'OTRO', 'nombre': 'Otro'},
]


def _coerce_string_list(value):
    if value is None:
        return []
    if isinstance(value, (list, tuple)):
        return [str(item).strip() for item in value if str(item).strip()]
    if isinstance(value, str):
        raw = value.strip()
        if not raw:
            return []
        try:
            parsed = json.loads(raw)
        except Exception:
            parsed = None
        if isinstance(parsed, (list, tuple)):
            return [str(item).strip() for item in parsed if str(item).strip()]
        return [raw]
    return [str(value).strip()]


def _coerce_json_payload(value, default):
    if value is None:
        return default
    if isinstance(value, (list, dict)):
        return value
    if isinstance(value, str):
        raw = value.strip()
        if not raw:
            return default
        try:
            return json.loads(raw)
        except Exception:
            return default
    return default


def _normalizar_grupo_requerimientos(value):
    parsed = _coerce_json_payload(value, [])
    if not isinstance(parsed, list):
        return []

    filas = []
    for item in parsed:
        if not isinstance(item, dict):
            continue
        fila = {
            'clave': (str(item.get('clave') or item.get('key') or '').strip().upper() or None),
            'nombre': (str(item.get('nombre') or item.get('label') or '').strip() or None),
            'seleccionado': bool(item.get('seleccionado')),
            'responsable': (str(item.get('responsable') or '').strip() or None),
            'celular': (str(item.get('celular') or '').strip() or None),
        }
        if any([fila['clave'], fila['nombre'], fila['seleccionado'], fila['responsable'], fila['celular']]):
            filas.append(fila)
    return filas


def _orden_caja_requiere_soporte(forma_pago, mixto_transferencia=None):
    forma = (forma_pago or '').strip().upper()
    if forma == 'TRANSFERENCIA':
        return True
    if forma != 'MIXTO':
        return False
    try:
        return Decimal(str(mixto_transferencia or 0)) > 0
    except Exception:
        return False


def _load_orden_caja_request_data():
    if request.is_json:
        return request.get_json(silent=True) or {}

    data = request.form.to_dict(flat=True)
    for field in (
        'enfasis',
        'paraclinicos',
        'laboratorio',
        'otros_servicios',
        'formas_autorizacion',
        'grupo_requerimientos',
    ):
        if field in request.form:
            data[field] = request.form.get(field)
    return data


def _serialize_orden_caja_adjunto(adjunto):
    return {
        'id': adjunto.id,
        'nombre_original': adjunto.nombre_original,
        'mime_type': adjunto.mime_type,
        'tamano_bytes': adjunto.tamano_bytes,
        'download_url': f'/api/comercial/caja/{adjunto.orden_id}/adjuntos/{adjunto.id}',
        'created_at': adjunto.created_at.strftime('%Y-%m-%d %H:%M:%S') if adjunto.created_at else None,
    }


def _guardar_adjunto_orden_caja(orden, archivo):
    if not archivo or not getattr(archivo, 'filename', ''):
        return

    nombre_original = secure_filename(os.path.basename(archivo.filename)) or f'adjunto_{uuid.uuid4().hex}'
    upload_root = current_app.config['UPLOAD_FOLDER']
    orden_dir = os.path.join(upload_root, 'comercial', 'caja', str(orden.id), 'transferencias')
    os.makedirs(orden_dir, exist_ok=True)

    nombre_guardado = f'{uuid.uuid4().hex}_{nombre_original}'
    ruta_absoluta = os.path.join(orden_dir, nombre_guardado)
    archivo.save(ruta_absoluta)

    db.session.add(OrdenServicioCajaAdjunto(
        orden_id=orden.id,
        nombre_original=nombre_original,
        ruta_relativa=os.path.relpath(ruta_absoluta, upload_root),
        mime_type=archivo.mimetype,
        tamano_bytes=os.path.getsize(ruta_absoluta) if os.path.exists(ruta_absoluta) else None,
    ))


def _eliminar_adjunto_orden_caja(adjunto):
    if not adjunto or not adjunto.ruta_relativa:
        return

    upload_root = os.path.abspath(current_app.config['UPLOAD_FOLDER'])
    ruta = os.path.abspath(os.path.join(upload_root, adjunto.ruta_relativa))
    if ruta.startswith(upload_root) and os.path.exists(ruta):
        try:
            os.remove(ruta)
        except OSError:
            logger.warning('No se pudo eliminar el adjunto de orden caja %s', ruta)


def _eliminar_adjuntos_orden_caja(orden):
    for adjunto in orden.adjuntos.all():
        _eliminar_adjunto_orden_caja(adjunto)
        db.session.delete(adjunto)


def _get_adjunto_orden_caja_path(adjunto):
    if not adjunto.ruta_relativa:
        raise FileNotFoundError('Adjunto no disponible')

    upload_root = os.path.abspath(current_app.config['UPLOAD_FOLDER'])
    ruta = os.path.abspath(os.path.join(upload_root, adjunto.ruta_relativa))
    if not ruta.startswith(upload_root):
        raise FileNotFoundError('Ruta de adjunto invalida')
    if not os.path.exists(ruta):
        raise FileNotFoundError('El archivo adjunto no existe en disco')
    return ruta


def _orden_caja_to_dict(o):
    """Serializa una OrdenServicioCaja a dict."""
    return {
        'id':                   o.id,
        'nro_orden':            o.nro_orden,
        'fecha_orden':          o.fecha_orden.strftime('%Y-%m-%d') if o.fecha_orden else None,
        'tipo_documento':       o.tipo_documento,
        'nro_documento':        o.nro_documento,
        'nombre_paciente':      o.nombre_paciente,
        'cargo_paciente':       o.cargo_paciente,
        'empresa':              o.empresa,
        'empresa_mision':       o.empresa_mision,
        'tipo_examen':          o.tipo_examen,
        'tipo_examen_otro':     o.tipo_examen_otro,
        'enfasis':              o.enfasis or [],
        'enfasis_otro':         o.enfasis_otro,
        'paraclinicos':         o.paraclinicos or [],
        'paraclinicos_otro':    o.paraclinicos_otro,
        'laboratorio':          o.laboratorio or [],
        'laboratorio_otro':     o.laboratorio_otro,
        'otros_servicios':      o.otros_servicios or [],
        'otros_servicios_otro': o.otros_servicios_otro,
        'total_costo':          float(o.total_costo or 0),
        'tipo_cliente':         o.tipo_cliente,
        'forma_pago':           o.forma_pago,
        'mixto_efectivo':       float(o.mixto_efectivo) if o.mixto_efectivo is not None else None,
        'mixto_transferencia':  float(o.mixto_transferencia) if o.mixto_transferencia is not None else None,
        'mixto_credito':        float(o.mixto_credito) if o.mixto_credito is not None else None,
        'formas_autorizacion':  o.formas_autorizacion or [],
        'autorizacion_observaciones': o.autorizacion_observaciones,
        'grupo_requerimientos': o.grupo_requerimientos or [],
        'numero_turno':         o.numero_turno,
        'estado':               o.estado,
        'motivo_anulacion':     o.motivo_anulacion,
        'observaciones':        o.observaciones,
        'cliente_id':           o.cliente_id,
        'adjuntos_transferencia': [
            _serialize_orden_caja_adjunto(adjunto)
            for adjunto in o.adjuntos.order_by(OrdenServicioCajaAdjunto.created_at.desc(), OrdenServicioCajaAdjunto.id.desc()).all()
        ],
        'usuario':              o.usuario.usuario if o.usuario else None,
        'created_at':           o.created_at.strftime('%Y-%m-%d %H:%M') if o.created_at else None,
        'updated_at':           o.updated_at.strftime('%Y-%m-%d %H:%M') if o.updated_at else None,
    }


def _parse_orden_caja_payload(data, *, es_nuevo=True, orden_actual=None):
    """Valida y construye el payload de una orden de caja."""
    errores = []

    nro_orden = (data.get('nro_orden') or '').strip()
    if not nro_orden:
        errores.append('El número de orden es obligatorio')

    fecha_str = (data.get('fecha_orden') or '').strip()
    fecha_orden = None
    if not fecha_str:
        errores.append('La fecha de la orden es obligatoria')
    else:
        try:
            fecha_orden = datetime.strptime(fecha_str, '%Y-%m-%d')
        except ValueError:
            errores.append('Formato de fecha inválido (YYYY-MM-DD)')

    tipo_doc = (data.get('tipo_documento') or '').strip().upper()
    if tipo_doc not in _TIPOS_DOCUMENTO_CAJA:
        errores.append('Tipo de documento debe ser CC, CE o PT')

    nro_doc = (data.get('nro_documento') or '').strip()
    if not nro_doc:
        errores.append('El número de documento es obligatorio')

    nombre = (data.get('nombre_paciente') or '').strip()
    if not nombre:
        errores.append('El nombre del paciente es obligatorio')

    forma_pago = (data.get('forma_pago') or '').strip().upper()
    if forma_pago not in _FORMAS_PAGO_CAJA:
        errores.append('Forma de pago inválida')

    try:
        total_costo = Decimal(str(data.get('total_costo') or 0))
    except Exception:
        total_costo = Decimal('0')
        errores.append('El total del costo es invalido')

    mixto_efectivo = mixto_transferencia = mixto_credito = None
    if forma_pago == 'MIXTO':
        try:
            mixto_efectivo     = Decimal(str(data.get('mixto_efectivo') or 0))
            mixto_transferencia = Decimal(str(data.get('mixto_transferencia') or 0))
            mixto_credito      = Decimal(str(data.get('mixto_credito') or 0))
            suma_mixto = mixto_efectivo + mixto_transferencia + mixto_credito
            if suma_mixto <= 0:
                errores.append('Para pago MIXTO debe indicar al menos un valor mayor a cero')
        except Exception:
            errores.append('Valores de pago MIXTO inválidos')

    formas_autorizacion = []
    for valor in _coerce_string_list(data.get('formas_autorizacion')):
        normalizado = valor.strip().upper()
        if normalizado in _OPCIONES_AUTORIZACION:
            formas_autorizacion.append(normalizado)

    grupo_requerimientos = _normalizar_grupo_requerimientos(data.get('grupo_requerimientos'))

    if errores:
        raise ValueError(' | '.join(errores))

    return {
        'nro_orden':            nro_orden,
        'fecha_orden':          fecha_orden,
        'tipo_documento':       tipo_doc,
        'nro_documento':        nro_doc,
        'nombre_paciente':      nombre,
        'cargo_paciente':       (data.get('cargo_paciente') or '').strip() or None,
        'empresa':              (data.get('empresa') or '').strip() or None,
        'empresa_mision':       (data.get('empresa_mision') or '').strip() or None,
        'tipo_examen':          (data.get('tipo_examen') or '').strip() or None,
        'tipo_examen_otro':     (data.get('tipo_examen_otro') or '').strip() or None,
        'enfasis':              _coerce_string_list(data.get('enfasis')),
        'enfasis_otro':         (data.get('enfasis_otro') or '').strip() or None,
        'paraclinicos':         _coerce_string_list(data.get('paraclinicos')),
        'paraclinicos_otro':    (data.get('paraclinicos_otro') or '').strip() or None,
        'laboratorio':          _coerce_string_list(data.get('laboratorio')),
        'laboratorio_otro':     (data.get('laboratorio_otro') or '').strip() or None,
        'otros_servicios':      _coerce_string_list(data.get('otros_servicios')),
        'otros_servicios_otro': (data.get('otros_servicios_otro') or '').strip() or None,
        'total_costo':          total_costo,
        'tipo_cliente':         (data.get('tipo_cliente') or '').strip().upper() or None,
        'forma_pago':           forma_pago,
        'mixto_efectivo':       mixto_efectivo,
        'mixto_transferencia':  mixto_transferencia,
        'mixto_credito':        mixto_credito,
        'formas_autorizacion':  formas_autorizacion,
        'autorizacion_observaciones': (data.get('autorizacion_observaciones') or '').strip() or None,
        'grupo_requerimientos': grupo_requerimientos,
        'numero_turno':         (data.get('numero_turno') or '').strip() or None,
        'observaciones':        (data.get('observaciones') or '').strip() or None,
        'cliente_id':           data.get('cliente_id') or None,
    }


# ---------------------------------------------------------------------------
# GET /api/comercial/caja/opciones  — catálogo de opciones del formulario
# ---------------------------------------------------------------------------
@comercial_bp.route('/caja/opciones', methods=['GET'])
@login_required
def opciones_orden_caja():
    return jsonify({
        'tipo_examen':   _OPCIONES_TIPO_EXAMEN,
        'enfasis':       _OPCIONES_ENFASIS,
        'paraclinicos':  _OPCIONES_PARACLINICOS,
        'laboratorio':   _OPCIONES_LABORATORIO,
        'otros_servicios': _OPCIONES_OTROS,
        'formas_autorizacion': _OPCIONES_AUTORIZACION,
        'grupos_requerimiento': _GRUPOS_REQUERIMIENTO,
        'tipos_documento': list(_TIPOS_DOCUMENTO_CAJA),
        'formas_pago':   list(_FORMAS_PAGO_CAJA),
        'tipos_cliente': list(_TIPOS_CLIENTE_CAJA),
        'estados':       list(_ESTADOS_ORDEN_CAJA),
    }), 200


# ---------------------------------------------------------------------------
# GET /api/comercial/caja  — consulta con filtros
# ---------------------------------------------------------------------------
@comercial_bp.route('/caja', methods=['GET'])
@login_required
def listar_ordenes_caja():
    try:
        _require_commercial_permission(PERMISO_CONSULTA_ATENCIONES)
    except PermissionError as exc:
        return jsonify({'error': str(exc)}), 403

    empresa    = (request.args.get('empresa') or '').strip()
    nro_orden  = (request.args.get('nro_orden') or '').strip()
    nro_doc    = (request.args.get('nro_documento') or '').strip()
    estado     = (request.args.get('estado') or '').strip().upper()
    fd_str     = (request.args.get('fecha_desde') or '').strip()
    fh_str     = (request.args.get('fecha_hasta') or '').strip()
    page       = max(1, int(request.args.get('page', 1)))
    per_page   = min(100, max(10, int(request.args.get('per_page', 50))))

    q = OrdenServicioCaja.query

    if empresa:
        q = q.filter(or_(
            OrdenServicioCaja.empresa.ilike(f'%{empresa}%'),
            OrdenServicioCaja.empresa_mision.ilike(f'%{empresa}%'),
        ))
    if nro_orden:
        q = q.filter(OrdenServicioCaja.nro_orden.ilike(f'%{nro_orden}%'))
    if nro_doc:
        q = q.filter(OrdenServicioCaja.nro_documento.ilike(f'%{nro_doc}%'))
    if estado and estado in _ESTADOS_ORDEN_CAJA:
        q = q.filter(OrdenServicioCaja.estado == estado)
    if fd_str:
        try:
            q = q.filter(OrdenServicioCaja.fecha_orden >= datetime.strptime(fd_str, '%Y-%m-%d'))
        except ValueError:
            pass
    if fh_str:
        try:
            q = q.filter(OrdenServicioCaja.fecha_orden <= datetime.strptime(f'{fh_str} 23:59:59', '%Y-%m-%d %H:%M:%S'))
        except ValueError:
            pass

    total   = q.count()
    ordenes = q.order_by(
        OrdenServicioCaja.fecha_orden.desc(),
        OrdenServicioCaja.nro_orden.asc(),
    ).offset((page - 1) * per_page).limit(per_page).all()

    return jsonify({
        'total':    total,
        'page':     page,
        'per_page': per_page,
        'pages':    (total + per_page - 1) // per_page,
        'ordenes':  [_orden_caja_to_dict(o) for o in ordenes],
    }), 200


# ---------------------------------------------------------------------------
# GET /api/comercial/caja/gaps  — detectar saltos en numeración
# ---------------------------------------------------------------------------
@comercial_bp.route('/caja/gaps', methods=['GET'])
@login_required
def detectar_gaps_ordenes_caja():
    """Detecta saltos en la numeración de órdenes para un rango de fechas."""
    try:
        _require_commercial_permission(PERMISO_CONSULTA_ATENCIONES)
    except PermissionError as exc:
        return jsonify({'error': str(exc)}), 403

    fd_str = (request.args.get('fecha_desde') or '').strip()
    fh_str = (request.args.get('fecha_hasta') or '').strip()

    q = OrdenServicioCaja.query.filter(
        OrdenServicioCaja.estado != 'ANULADO'
    )
    if fd_str:
        try:
            q = q.filter(OrdenServicioCaja.fecha_orden >= datetime.strptime(fd_str, '%Y-%m-%d'))
        except ValueError:
            pass
    if fh_str:
        try:
            q = q.filter(OrdenServicioCaja.fecha_orden <= datetime.strptime(f'{fh_str} 23:59:59', '%Y-%m-%d %H:%M:%S'))
        except ValueError:
            pass

    ordenes = q.order_by(OrdenServicioCaja.nro_orden.asc()).all()

    # Extraer números de las órdenes (solo la parte numérica)
    import re as _re
    numeros = []
    for o in ordenes:
        m = _re.search(r'\d+', o.nro_orden or '')
        if m:
            numeros.append((int(m.group()), o.nro_orden))

    numeros.sort(key=lambda x: x[0])
    gaps = []
    for i in range(1, len(numeros)):
        prev_num, prev_nro = numeros[i - 1]
        curr_num, curr_nro = numeros[i]
        if curr_num - prev_num > 1:
            faltantes = list(range(prev_num + 1, curr_num))
            gaps.append({
                'entre':     f'{prev_nro} y {curr_nro}',
                'faltantes': faltantes[:20],  # máximo 20 por gap
                'cantidad':  len(faltantes),
            })

    return jsonify({
        'total_ordenes': len(ordenes),
        'gaps':          gaps,
        'tiene_gaps':    len(gaps) > 0,
    }), 200


# ---------------------------------------------------------------------------
# POST /api/comercial/caja  — crear orden
# ---------------------------------------------------------------------------
@comercial_bp.route('/caja', methods=['POST'])
@login_required
def crear_orden_caja():
    try:
        _require_commercial_permission(PERMISO_CARGUE_ATENCIONES)
    except PermissionError as exc:
        return jsonify({'error': str(exc)}), 403

    data = _load_orden_caja_request_data()
    archivos = [archivo for archivo in request.files.getlist('transferencia_adjuntos') if getattr(archivo, 'filename', '')]
    try:
        payload = _parse_orden_caja_payload(data)
    except ValueError as exc:
        return jsonify({'error': str(exc)}), 400

    if _orden_caja_requiere_soporte(payload['forma_pago'], payload['mixto_transferencia']) and not archivos:
        return jsonify({'error': 'Debes adjuntar al menos un recibo cuando la orden incluye transferencia'}), 400

    orden = OrdenServicioCaja(
        **payload,
        estado     = 'INGRESADO',
        usuario_id = current_user.id,
    )
    db.session.add(orden)
    try:
        db.session.flush()
        for archivo in archivos:
            _guardar_adjunto_orden_caja(orden, archivo)
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        logger.error('Error creando orden caja: %s', exc)
        return jsonify({'error': 'No se pudo guardar la orden'}), 500

    return jsonify({'orden': _orden_caja_to_dict(orden)}), 201


# ---------------------------------------------------------------------------
# GET /api/comercial/caja/<id>  — detalle
# ---------------------------------------------------------------------------
@comercial_bp.route('/caja/<int:orden_id>', methods=['GET'])
@login_required
def obtener_orden_caja(orden_id):
    try:
        _require_commercial_permission(PERMISO_CONSULTA_ATENCIONES)
    except PermissionError as exc:
        return jsonify({'error': str(exc)}), 403

    orden = OrdenServicioCaja.query.get_or_404(orden_id)
    return jsonify({'orden': _orden_caja_to_dict(orden)}), 200


# ---------------------------------------------------------------------------
# PUT /api/comercial/caja/<id>  — editar (solo INGRESADO)
# ---------------------------------------------------------------------------
@comercial_bp.route('/caja/<int:orden_id>', methods=['PUT'])
@login_required
def editar_orden_caja(orden_id):
    try:
        _require_commercial_permission(PERMISO_CARGUE_ATENCIONES)
    except PermissionError as exc:
        return jsonify({'error': str(exc)}), 403

    orden = OrdenServicioCaja.query.get_or_404(orden_id)
    if orden.estado != 'INGRESADO':
        return jsonify({'error': f'Solo se pueden editar órdenes en estado INGRESADO. Esta está en {orden.estado}'}), 409

    data = _load_orden_caja_request_data()
    archivos = [archivo for archivo in request.files.getlist('transferencia_adjuntos') if getattr(archivo, 'filename', '')]
    try:
        payload = _parse_orden_caja_payload(data, es_nuevo=False, orden_actual=orden)
    except ValueError as exc:
        return jsonify({'error': str(exc)}), 400

    for campo, valor in payload.items():
        setattr(orden, campo, valor)

    requiere_soporte = _orden_caja_requiere_soporte(payload['forma_pago'], payload['mixto_transferencia'])
    if requiere_soporte and not archivos and not orden.adjuntos.count():
        return jsonify({'error': 'Debes adjuntar al menos un recibo cuando la orden incluye transferencia'}), 400

    try:
        if not requiere_soporte:
            _eliminar_adjuntos_orden_caja(orden)
        else:
            for archivo in archivos:
                _guardar_adjunto_orden_caja(orden, archivo)
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        return jsonify({'error': 'No se pudo actualizar la orden'}), 500

    return jsonify({'orden': _orden_caja_to_dict(orden)}), 200


# ---------------------------------------------------------------------------
# POST /api/comercial/caja/<id>/cambiar-estado  — aprobar / terminar / anular
# ---------------------------------------------------------------------------
@comercial_bp.route('/caja/<int:orden_id>/cambiar-estado', methods=['POST'])
@login_required
def cambiar_estado_orden_caja(orden_id):
    try:
        _require_commercial_permission(PERMISO_CARGUE_ATENCIONES)
    except PermissionError as exc:
        return jsonify({'error': str(exc)}), 403

    orden  = OrdenServicioCaja.query.get_or_404(orden_id)
    data   = request.get_json() or {}
    nuevo  = (data.get('estado') or '').strip().upper()

    if nuevo not in _ESTADOS_ORDEN_CAJA:
        return jsonify({'error': f'Estado inválido. Use: {", ".join(_ESTADOS_ORDEN_CAJA)}'}), 400

    # Validar transiciones permitidas
    transiciones = {
        'INGRESADO': {'APROBADO', 'ANULADO'},
        'APROBADO':  {'TERMINADO', 'ANULADO'},
        'TERMINADO': set(),
        'ANULADO':   set(),
    }
    if nuevo not in transiciones.get(orden.estado, set()):
        return jsonify({'error': f'No se puede pasar de {orden.estado} a {nuevo}'}), 409

    if nuevo == 'ANULADO':
        motivo = (data.get('motivo') or '').strip()
        if not motivo:
            return jsonify({'error': 'El motivo de anulación es obligatorio'}), 400
        orden.motivo_anulacion  = motivo
        orden.usuario_anula_id  = current_user.id
        orden.fecha_anulacion   = datetime.utcnow()

    elif nuevo == 'APROBADO':
        orden.usuario_aprueba_id = current_user.id
        orden.fecha_aprobacion   = datetime.utcnow()

    elif nuevo == 'TERMINADO':
        orden.usuario_termina_id = current_user.id
        orden.fecha_terminacion  = datetime.utcnow()

    orden.estado = nuevo

    try:
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        return jsonify({'error': 'No se pudo cambiar el estado'}), 500

    return jsonify({'orden': _orden_caja_to_dict(orden)}), 200


@comercial_bp.route('/caja/<int:orden_id>/adjuntos/<int:adjunto_id>', methods=['GET'])
@login_required
def descargar_adjunto_orden_caja(orden_id, adjunto_id):
    try:
        _require_commercial_permission(PERMISO_CONSULTA_ATENCIONES)
    except PermissionError as exc:
        return jsonify({'error': str(exc)}), 403

    orden = OrdenServicioCaja.query.get_or_404(orden_id)
    adjunto = OrdenServicioCajaAdjunto.query.filter_by(id=adjunto_id, orden_id=orden.id).first_or_404()
    try:
        ruta = _get_adjunto_orden_caja_path(adjunto)
        return send_file(ruta, as_attachment=True, download_name=adjunto.nombre_original)
    except FileNotFoundError as exc:
        return jsonify({'error': str(exc)}), 404
    except Exception as exc:
        logger.error('Error descargando adjunto de orden caja %s: %s', adjunto_id, exc)
        return jsonify({'error': 'No se pudo descargar el adjunto'}), 500


# ---------------------------------------------------------------------------
# DELETE /api/comercial/caja/<id>  — eliminar (solo INGRESADO, solo admin)
# ---------------------------------------------------------------------------
@comercial_bp.route('/caja/<int:orden_id>', methods=['DELETE'])
@login_required
def eliminar_orden_caja(orden_id):
    if not _is_admin_user():
        return jsonify({'error': 'Solo el administrador puede eliminar órdenes'}), 403

    orden = OrdenServicioCaja.query.get_or_404(orden_id)
    if orden.estado != 'INGRESADO':
        return jsonify({'error': 'Solo se pueden eliminar órdenes en estado INGRESADO'}), 409

    try:
        _eliminar_adjuntos_orden_caja(orden)
        db.session.delete(orden)
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        return jsonify({'error': 'No se pudo eliminar la orden'}), 500

    return jsonify({'ok': True}), 200
